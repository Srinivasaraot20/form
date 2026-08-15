from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Count
from registrations.models import StudentApplication, UploadedDocument, SUBMITTED_STATUS_CODES
from dashboard.dashboard_data import DASHBOARD_STATUS_CODES, DRAFT_CODE
from masterdata.models import State, District, Program, TrainingPartner, Community, Occupation, Qualification, Religion, VerificationStatus, ApplicationStatus, BatchCode
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
from datetime import timedelta

@login_required
@user_passes_test(lambda u: u.is_staff)
def dashboard_home(request):
    """
    Renders the main dashboard with live ORM data metrics.
    Every number is calculated on the fly from the source tables by the
    dashboard_data service - nothing is hard-coded.
    """
    from .dashboard_data import get_dashboard_data, DASHBOARD_STATUS_CODES
    data = get_dashboard_data(request)
    context = dict(data)
    context['recent_apps'] = StudentApplication.objects.filter(
        status__code__in=DASHBOARD_STATUS_CODES
    ).select_related('program_opting', 'status').order_by('-submission_date')[:5]
    return render(request, 'dashboard/index.html', context)


# ==========================================
# PLACEHOLDER VIEWS TO ELIMINATE 404s
# ==========================================

def _name(model, pk):
    """Resolve a primary key to a human-readable name for filter chips."""
    if not pk:
        return ''
    obj = model.objects.filter(pk=pk).first()
    return str(obj) if obj else ''


def _status_name(code):
    """Resolve a status/verification code to its human-readable name for chips."""
    if not code:
        return ''
    if code == 'NONE':
        return 'Not Verified'
    obj = VerificationStatus.objects.filter(code=code).first()
    if not obj:
        obj = ApplicationStatus.objects.filter(code=code).first()
    return str(obj) if obj else ''


def _quality_label(quality):
    """Human-readable label for a data-quality drill-down filter."""
    return {
        'missing_docs': 'Missing Documents',
        'invalid_email': 'Invalid Email',
        'incomplete_education': 'Incomplete Education Data',
        'dup_mobile': 'Duplicate Mobile',
        'dup_aadhaar': 'Duplicate Aadhaar',
        'dup_email': 'Duplicate Email',
    }.get(quality or '', '')

def _filtered_applications(request, base=None):
    """
    Applies every supported Applications filter to a base queryset and returns it.
    Shared by the Applications page and the Export CSV endpoint so exports always
    respect the same active filters. Filtering is purely READ-ONLY queryset logic.
    """
    qs = base if base is not None else StudentApplication.objects.all()
    
    # Search (App No / Name / Aadhaar / Mobile / Email)
    search_query = request.GET.get('search', '').strip()
    if search_query:
        qs = qs.filter(
            Q(application_number__icontains=search_query) |
            Q(full_name__icontains=search_query) |
            Q(aadhaar_number__icontains=search_query) |
            Q(mobile_number__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    # Location
    state_id = request.GET.get('state')
    district_id = request.GET.get('district')
    if state_id:
        qs = qs.filter(state_id=state_id)
    if district_id:
        qs = qs.filter(district_id=district_id)
    
    # Program / Qualification / Training Partner
    program_id = request.GET.get('program')
    qual_id = request.GET.get('applying_qualification')
    qualification_id = request.GET.get('qualification')
    partner_id = request.GET.get('training_partner')
    if program_id:
        qs = qs.filter(program_opting_id=program_id)
    if qual_id:
        qs = qs.filter(applying_qualification=qual_id)
    if qualification_id:
        qs = qs.filter(qualification_id=qualification_id)
    if partner_id:
        qs = qs.filter(training_partner_id=partner_id)
    
    # Application status: the workflow supports exactly two UI-level buckets.
    # "draft" (INCOMPLETE) and "submitted" (every officially-submitted status,
    # i.e. registrations.SUBMITTED_STATUS_CODES - this includes the current
    # SUBMITTED code and the legacy PENDING/APPROVED/REJECTED/CORRECTION codes
    # that existing records use, so pre-existing data always appears here).
    # Bucket membership is based on the record's actual status ForeignKey -
    # never on submission_date or on whether other fields are populated.
    # The selected bucket always ANDs with every other filter.
    status_id = request.GET.get('status')
    if status_id == 'draft':
        qs = qs.filter(status__code=DRAFT_CODE)
    elif status_id == 'submitted':
        qs = qs.filter(status__code__in=SUBMITTED_STATUS_CODES)
    else:
        # "All Statuses": show both Submitted + Draft buckets.
        qs = qs.filter(status__code__in=DASHBOARD_STATUS_CODES)
    
    # Gender / Category (community) / Occupation
    gender = request.GET.get('gender')
    category_id = request.GET.get('category')
    occupation_id = request.GET.get('occupation')
    if gender:
        qs = qs.filter(gender=gender)
    if category_id:
        qs = qs.filter(community_id=category_id)
    if occupation_id:
        qs = qs.filter(occupation_id=occupation_id)
    
    # Year of passing
    year = request.GET.get('year_of_passing')
    if year:
        qs = qs.filter(year_of_passing=year)
    
    # Submission date: quick range OR custom from/to. Custom wins.
    today = timezone.localdate()
    date_range = request.GET.get('date_range')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_range == 'today':
        qs = qs.filter(submission_date__date=today)
    elif date_range == 'yesterday':
        qs = qs.filter(submission_date__date=today - timedelta(days=1))
    elif date_range == 'last7':
        qs = qs.filter(submission_date__date__gte=today - timedelta(days=6))
    elif date_range == 'last30':
        qs = qs.filter(submission_date__date__gte=today - timedelta(days=29))
    elif date_range == 'this_month':
        qs = qs.filter(submission_date__date__year=today.year, submission_date__date__month=today.month)
    elif date_from:
        try:
            d_from = timezone.datetime.strptime(date_from, '%Y-%m-%d').date()
            qs = qs.filter(submission_date__date__gte=d_from)
        except ValueError:
            pass
        if date_to:
            try:
                d_to = timezone.datetime.strptime(date_to, '%Y-%m-%d').date()
                qs = qs.filter(submission_date__date__lte=d_to)
            except ValueError:
                pass
    
    # Data-quality drill-downs (used by dashboard alert cards)
    quality = request.GET.get('quality')
    if quality:
        from .dashboard_data import _ids_for_quality
        qs = qs.filter(id__in=_ids_for_quality(quality))
    
    return qs


@login_required
@user_passes_test(lambda u: u.is_staff)
def application_list(request):
    """
    Renders the Applications table with full search/filtering/pagination.
    All filters work against the database queryset (AND logic) and are
    preserved in URL query parameters so refresh/back/forward keep them.
    """
    applications = _filtered_applications(request).select_related(
        'program_opting', 'status', 'verification_status', 'state', 'district',
        'qualification', 'community', 'occupation'
    ).order_by('-submission_date')
    
    # Result count (used for the "Showing N applications" summary)
    total_count = applications.count()
    
    # Pagination (filters preserved via query string built below)
    paginator = Paginator(applications, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Preserve all current filters across pagination links
    preserve = {
        'search': request.GET.get('search', ''),
        'state': request.GET.get('state', ''),
        'district': request.GET.get('district', ''),
        'program': request.GET.get('program', ''),
        'applying_qualification': request.GET.get('applying_qualification', ''),
        'qualification': request.GET.get('qualification', ''),
        'training_partner': request.GET.get('training_partner', ''),
        'status': request.GET.get('status', ''),
        'gender': request.GET.get('gender', ''),
        'category': request.GET.get('category', ''),
        'occupation': request.GET.get('occupation', ''),
        'year_of_passing': request.GET.get('year_of_passing', ''),
        'date_range': request.GET.get('date_range', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
        'quality': request.GET.get('quality', ''),
    }
    query_string = '&'.join(f'{k}={v}' for k, v in preserve.items() if v)
    
    # Years descending, from existing records only
    years = (
        StudentApplication.objects.order_by()
        .exclude(year_of_passing__isnull=True)
        .values_list('year_of_passing', flat=True)
        .distinct()
    )
    years = sorted(set(years), reverse=True)

    context = {
        'title': 'All Applications',
        'page_obj': page_obj,
        'total_count': total_count,
        'query_string': query_string,
        'states': State.objects.filter(is_active=True).order_by('name'),
        'districts': District.objects.filter(is_active=True).order_by('name'),
        'programs': Program.objects.filter(is_active=True),
        'applying_qualifications': StudentApplication.APPLYING_QUALIFICATION_CHOICES,
        'training_partners': TrainingPartner.objects.filter(is_active=True),
        'statuses': [
            {'value': 'draft', 'name': 'Draft'},
            {'value': 'submitted', 'name': 'Submitted'},
        ],
        'genders': StudentApplication.GENDER_CHOICES,
        'categories': Community.objects.filter(is_active=True),
        'occupations': Occupation.objects.filter(is_active=True),
        'years': years,
        # Pass current filters to preserve in UI
        'current_search': request.GET.get('search', ''),
        'current_state': request.GET.get('state', ''),
        'current_district': request.GET.get('district', ''),
        'current_program': request.GET.get('program', ''),
        'current_applying_qualification': request.GET.get('applying_qualification', ''),
        'current_qualification': request.GET.get('qualification', ''),
        'current_training_partner': request.GET.get('training_partner', ''),
        'current_status': request.GET.get('status', ''),
        'current_gender': request.GET.get('gender', ''),
        'current_category': request.GET.get('category', ''),
        'current_occupation': request.GET.get('occupation', ''),
        'current_year': request.GET.get('year_of_passing', ''),
        'current_date_range': request.GET.get('date_range', ''),
        'current_date_from': request.GET.get('date_from', ''),
        'current_date_to': request.GET.get('date_to', ''),
        'current_quality': request.GET.get('quality', ''),
        'current_quality_label': _quality_label(request.GET.get('quality', '')),
        # Display names for Active Filter chips
        'current_state_name': _name(State, request.GET.get('state')),
        'current_district_name': _name(District, request.GET.get('district')),
        'current_program_name': _name(Program, request.GET.get('program')),
        'current_qualification_name': _name(Qualification, request.GET.get('qualification')),
        'current_category_name': _name(Community, request.GET.get('category')),
        'current_occupation_name': _name(Occupation, request.GET.get('occupation')),
        'current_training_partner_name': _name(TrainingPartner, request.GET.get('training_partner')),
    }
    return render(request, 'dashboard/applications.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def application_pending(request):
    """
    Dedicated Pending Applications workspace.

    Shows ONLY officially submitted applications whose current application
    status is PENDING (Pending Review). This uses the SAME status code and
    queryset logic as the dashboard "Pending" count, so both always agree.
    Draft/auto-save (INCOMPLETE) and Preview records never appear here.

    Includes: summary counts, search, verification/date/location filters,
    sorting, bulk selection/actions, and per-row verification actions.
    """
    # The dashboard Pending count uses: submitted_apps.filter(status__code='PENDING')
    # We mirror it exactly. Only final-submitted PENDING records appear.
    base = StudentApplication.objects.filter(status__code='PENDING')

    today = timezone.localdate()
    week_start = today - timedelta(days=today.weekday())

    # ---- Summary counts (unfiltered, always reflect the full pending set) ----
    total_pending = base.count()
    pending_today = base.filter(submission_date__date=today).count()
    pending_week = base.filter(submission_date__date__gte=week_start).count()
    oldest_pending = base.order_by('submission_date').first()

    # ---- Search: App No / Name / Aadhaar / Mobile / Email ----
    search_query = request.GET.get('search', '').strip()
    if search_query:
        base = base.filter(
            Q(application_number__icontains=search_query) |
            Q(full_name__icontains=search_query) |
            Q(aadhaar_number__icontains=search_query) |
            Q(mobile_number__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    # ---- Application Status ----
    # This page shows ONLY PENDING (Pending Review) records by design, so the
    # status filter's only meaningful option here is PENDING. Approved/Rejected
    # and Draft are never offered and never appear on this page.
    app_status = request.GET.get('application_status')
    if app_status == 'PENDING':
        base = base.filter(status__code='PENDING')
    elif app_status and app_status != 'all':
        base = base.filter(status__code=app_status)

    # ---- Verification status (PENDING/VERIFIED/CORRECTION) ----
    v_status = request.GET.get('verification_status')
    if v_status == 'PENDING':
        base = base.filter(Q(verification_status__isnull=True) | Q(verification_status__code='PENDING'))
    elif v_status:
        base = base.filter(verification_status__code=v_status)

    # ---- Submission date quick ranges + custom range ----
    today = timezone.localdate()
    date_range = request.GET.get('date_range')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_range == 'today':
        base = base.filter(submission_date__date=today)
    elif date_range == 'yesterday':
        base = base.filter(submission_date__date=today - timedelta(days=1))
    elif date_range == 'this_week':
        base = base.filter(submission_date__date__gte=week_start)
    elif date_range == 'this_month':
        base = base.filter(submission_date__date__year=today.year, submission_date__date__month=today.month)
    elif date_range == 'last7':
        base = base.filter(submission_date__date__gte=today - timedelta(days=6))
    elif date_range == 'last30':
        base = base.filter(submission_date__date__gte=today - timedelta(days=29))
    elif date_from:
        try:
            base = base.filter(submission_date__date__gte=timezone.datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
        if date_to:
            try:
                base = base.filter(submission_date__date__lte=timezone.datetime.strptime(date_to, '%Y-%m-%d').date())
            except ValueError:
                pass

    # ---- Location / Program / Qualification / Partner / Gender ----
    state_id = request.GET.get('state')
    district_id = request.GET.get('district')
    program_id = request.GET.get('program')
    qualification_id = request.GET.get('qualification')
    partner_id = request.GET.get('training_partner')
    gender = request.GET.get('gender')
    if state_id:
        base = base.filter(state_id=state_id)
    if district_id:
        base = base.filter(district_id=district_id)
    if program_id:
        base = base.filter(program_opting_id=program_id)
    if qualification_id:
        base = base.filter(qualification_id=qualification_id)
    if partner_id:
        base = base.filter(training_partner_id=partner_id)
    if gender:
        base = base.filter(gender=gender)

    # ---- Sorting (default: oldest pending first) ----
    sort = request.GET.get('sort', 'oldest')
    sort_map = {
        'newest': '-submission_date',
        'oldest': 'submission_date',
        'name_asc': 'full_name',
        'name_desc': '-full_name',
        'app_no_asc': 'application_number',
        'app_no_desc': '-application_number',
    }
    order_by = sort_map.get(sort, 'submission_date')
    applications = base.select_related(
        'program_opting', 'status', 'verification_status', 'state', 'district',
        'qualification', 'training_partner'
    ).prefetch_related('documents').order_by(order_by)

    # ---- Result count + pagination (filters preserved via query string) ----
    total_count = applications.count()
    # True only when the user is actively filtering (for "Showing X of Y")
    filter_params = {
        'search': request.GET.get('search', ''),
        'verification_status': request.GET.get('verification_status', ''),
        'date_range': request.GET.get('date_range', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
        'state': request.GET.get('state', ''),
        'district': request.GET.get('district', ''),
        'program': request.GET.get('program', ''),
        'qualification': request.GET.get('qualification', ''),
        'training_partner': request.GET.get('training_partner', ''),
        'gender': request.GET.get('gender', ''),
    }
    filters_active = any(v for v in filter_params.values())
    paginator = Paginator(applications, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Attach photo URL to each app for this page (avoids per-row doc scans)
    for app in page_obj.object_list:
        photo = next(
            (d for d in app.documents.all() if 'passport' in d.doc_type.lower() and d.compressed_file),
            None
        )
        app.photo_url = photo.compressed_file.url if photo else None

    preserve = {
        'search': request.GET.get('search', ''),
        'application_status': request.GET.get('application_status', ''),
        'verification_status': request.GET.get('verification_status', ''),
        'date_range': request.GET.get('date_range', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
        'state': request.GET.get('state', ''),
        'district': request.GET.get('district', ''),
        'program': request.GET.get('program', ''),
        'qualification': request.GET.get('qualification', ''),
        'training_partner': request.GET.get('training_partner', ''),
        'gender': request.GET.get('gender', ''),
        'sort': request.GET.get('sort', ''),
    }
    query_string = '&'.join(f'{k}={v}' for k, v in preserve.items() if v)

    context = {
        'title': 'Pending Applications',
        'page_obj': page_obj,
        'total_count': total_count,
        'filters_active': filters_active,
        'total_pending': total_pending,
        'pending_today': pending_today,
        'pending_week': pending_week,
        'oldest_pending': oldest_pending,
        'query_string': query_string,
        'states': State.objects.filter(is_active=True).order_by('name'),
        'districts': District.objects.filter(is_active=True).order_by('name'),
        'programs': Program.objects.filter(is_active=True),
        'qualifications': Qualification.objects.filter(is_active=True),
        'training_partners': TrainingPartner.objects.filter(is_active=True),
        'genders': StudentApplication.GENDER_CHOICES,
        'verification_statuses': VerificationStatus.objects.filter(is_active=True),
        'current_search': request.GET.get('search', ''),
        'current_application_status': request.GET.get('application_status', 'PENDING'),
        'current_verification_status': request.GET.get('verification_status', ''),
        'current_date_range': request.GET.get('date_range', ''),
        'current_date_from': request.GET.get('date_from', ''),
        'current_date_to': request.GET.get('date_to', ''),
        'current_state': request.GET.get('state', ''),
        'current_district': request.GET.get('district', ''),
        'current_program': request.GET.get('program', ''),
        'current_qualification': request.GET.get('qualification', ''),
        'current_training_partner': request.GET.get('training_partner', ''),
        'current_gender': request.GET.get('gender', ''),
        'current_sort': request.GET.get('sort', 'oldest'),
        # Display names for filter chips
        'current_state_name': _name(State, request.GET.get('state')),
        'current_district_name': _name(District, request.GET.get('district')),
        'current_program_name': _name(Program, request.GET.get('program')),
        'current_qualification_name': _name(Qualification, request.GET.get('qualification')),
        'current_training_partner_name': _name(TrainingPartner, request.GET.get('training_partner')),
    }
    return render(request, 'dashboard/pending.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def application_new(request):
    """
    "New (Last 24h)" applications workspace.

    Shows ONLY StudentApplication records created within the last 24 hours,
    using the application's actual creation timestamp (submission_date,
    auto_now_add). The 24h window is recomputed from timezone.now() on every
    request, so this page always reflects the latest database state.

    Read-only: queries existing records only. Never creates, updates, or
    duplicates applications. Includes search, program/qualification/partner/
    gender/status filters, sorting, bulk selection, and a filter-aware empty
    state.
    """
    now = timezone.now()
    since_24h = now - timedelta(hours=24)
    base = StudentApplication.objects.filter(submission_date__gte=since_24h)

    # ---- Summary ----
    total_new = base.count()
    today = timezone.localdate()

    # ---- Search: App No / Name / Aadhaar / Mobile / Email ----
    search_query = request.GET.get('search', '').strip()
    if search_query:
        base = base.filter(
            Q(application_number__icontains=search_query) |
            Q(full_name__icontains=search_query) |
            Q(aadhaar_number__icontains=search_query) |
            Q(mobile_number__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    # ---- Application Status (Submitted / Draft only, no Approve/Reject) ----
    # Same two UI-level buckets as the Applications page. The underlying
    # database statuses are never modified.
    app_status = request.GET.get('application_status')
    if app_status == 'draft':
        base = base.filter(status__code='INCOMPLETE')
    elif app_status == 'submitted':
        base = base.filter(status__code__in=SUBMITTED_STATUS_CODES)

    # ---- Program / Qualification / Partner / Gender ----
    program_id = request.GET.get('program')
    qualification_id = request.GET.get('qualification')
    partner_id = request.GET.get('training_partner')
    gender = request.GET.get('gender')
    if program_id:
        base = base.filter(program_opting_id=program_id)
    if qualification_id:
        base = base.filter(qualification_id=qualification_id)
    if partner_id:
        base = base.filter(training_partner_id=partner_id)
    if gender:
        base = base.filter(gender=gender)

    # ---- Sorting (default: newest first) ----
    sort = request.GET.get('sort', 'newest')
    sort_map = {
        'newest': '-submission_date',
        'oldest': 'submission_date',
        'name_asc': 'full_name',
        'name_desc': '-full_name',
    }
    order_by = sort_map.get(sort, '-submission_date')
    applications = base.select_related(
        'program_opting', 'status', 'verification_status', 'state', 'district',
        'qualification', 'training_partner'
    ).prefetch_related('documents').order_by(order_by)

    # ---- Result count + pagination ----
    total_count = applications.count()
    filter_params = {
        'search': request.GET.get('search', ''),
        'application_status': request.GET.get('application_status', ''),
        'program': request.GET.get('program', ''),
        'qualification': request.GET.get('qualification', ''),
        'training_partner': request.GET.get('training_partner', ''),
        'gender': request.GET.get('gender', ''),
    }
    filters_active = any(v for v in filter_params.values())
    paginator = Paginator(applications, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Attach photo URL to each app for this page (avoids per-row doc scans)
    for app in page_obj.object_list:
        photo = next(
            (d for d in app.documents.all() if 'passport' in d.doc_type.lower() and d.compressed_file),
            None
        )
        app.photo_url = photo.compressed_file.url if photo else None

    preserve = {
        'search': request.GET.get('search', ''),
        'application_status': request.GET.get('application_status', ''),
        'program': request.GET.get('program', ''),
        'qualification': request.GET.get('qualification', ''),
        'training_partner': request.GET.get('training_partner', ''),
        'gender': request.GET.get('gender', ''),
        'sort': request.GET.get('sort', ''),
    }
    query_string = '&'.join(f'{k}={v}' for k, v in preserve.items() if v)

    context = {
        'title': 'New Applications — Last 24 Hours',
        'page_obj': page_obj,
        'total_count': total_count,
        'filters_active': filters_active,
        'total_new': total_new,
        'query_string': query_string,
        'since_24h': since_24h,
        'programs': Program.objects.filter(is_active=True),
        'qualifications': Qualification.objects.filter(is_active=True),
        'training_partners': TrainingPartner.objects.filter(is_active=True),
        'genders': StudentApplication.GENDER_CHOICES,
        'statuses': [
            {'value': 'draft', 'name': 'Draft'},
            {'value': 'submitted', 'name': 'Submitted'},
        ],
        'current_search': request.GET.get('search', ''),
        'current_application_status': request.GET.get('application_status', ''),
        'current_program': request.GET.get('program', ''),
        'current_qualification': request.GET.get('qualification', ''),
        'current_training_partner': request.GET.get('training_partner', ''),
        'current_gender': request.GET.get('gender', ''),
        'current_sort': request.GET.get('sort', 'newest'),
        # Display names for filter chips
        'current_program_name': _name(Program, request.GET.get('program')),
        'current_qualification_name': _name(Qualification, request.GET.get('qualification')),
        'current_training_partner_name': _name(TrainingPartner, request.GET.get('training_partner')),
    }
    return render(request, 'dashboard/new_applications.html', context)

# =====================================================================
# STUDENTS MODULE
# =====================================================================

def _attach_photo_urls(apps):
    """Attach app.photo_url (passport compressed file) to each app in a page."""
    for app in apps:
        photo = next(
            (d for d in app.documents.all() if 'passport' in d.doc_type.lower() and d.compressed_file),
            None
        )
        app.photo_url = photo.compressed_file.url if photo else None
    return apps


@login_required
@user_passes_test(lambda u: u.is_staff)
def student_directory(request):
    """
    Student Directory: student-centric management of all applicants.
    Unlike the Applications workflow pages, this is a people-focused view —
    every row is a student (their application record) with profile access,
    documents, timeline and edit links. Read-only: never creates records.
    """
    base = StudentApplication.objects.all()

    # ---- Statistics (always reflect the full student population) ----
    total_students = base.count()
    submitted_count = base.filter(status__code__in=SUBMITTED_STATUS_CODES).count()
    draft_count = base.filter(status__code='INCOMPLETE').count()
    verified_count = base.filter(verification_status__code='VERIFIED').count()
    correction_count = base.filter(verification_status__code='CORRECTION').count()
    # "Pending Verification" = officially submitted applications that have not yet
    # been verified or marked for correction.
    pending_verification_count = (
        base.filter(status__code__in=SUBMITTED_STATUS_CODES)
        .exclude(verification_status__code__in=['VERIFIED', 'CORRECTION'])
        .count()
    )

    # ---- Search: App No / Name / Mobile / Aadhaar / Email / ABC-APAAR / Reg No ----
    search_query = request.GET.get('search', '').strip()
    if search_query:
        base = base.filter(
            Q(application_number__icontains=search_query) |
            Q(full_name__icontains=search_query) |
            Q(mobile_number__icontains=search_query) |
            Q(aadhaar_number__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(abc_id__icontains=search_query) |
            Q(registration_number__icontains=search_query)
        )

    # ---- Filters ----
    program_id = request.GET.get('program')
    qualification_id = request.GET.get('qualification')
    gender = request.GET.get('gender')
    state_id = request.GET.get('state')
    district_id = request.GET.get('district')
    institution = request.GET.get('institution', '').strip()
    partner_id = request.GET.get('training_partner')
    app_status_code = request.GET.get('app_status', '').strip()
    verification_code = request.GET.get('verification', '').strip()
    passing_year = request.GET.get('year_of_passing', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    if program_id:
        base = base.filter(program_opting_id=program_id)
    if qualification_id:
        base = base.filter(qualification_id=qualification_id)
    if gender:
        base = base.filter(gender=gender)
    if state_id:
        base = base.filter(state_id=state_id)
    if district_id:
        base = base.filter(district_id=district_id)
    if institution:
        base = base.filter(institution__icontains=institution)
    if partner_id:
        base = base.filter(training_partner_id=partner_id)
    if app_status_code:
        base = base.filter(status__code=app_status_code)
    if verification_code:
        if verification_code == 'NONE':
            # "Not Verified" = no VERIFIED/CORRECTION verification status
            # (includes records with a pending/null verification status).
            base = base.exclude(verification_status__code__in=['VERIFIED', 'CORRECTION'])
        else:
            base = base.filter(verification_status__code=verification_code)
    if passing_year:
        base = base.filter(year_of_passing=passing_year)
    if date_from:
        base = base.filter(submission_date__date__gte=date_from)
    if date_to:
        base = base.filter(submission_date__date__lte=date_to)

    # ---- Sorting ----
    sort = request.GET.get('sort', 'registration_desc')
    sort_map = {
        'name_asc': 'full_name',
        'name_desc': '-full_name',
        'registration_asc': 'submission_date',
        'registration_desc': '-submission_date',
        'app_no_asc': 'application_number',
        'app_no_desc': '-application_number',
    }
    order_by = sort_map.get(sort, '-submission_date')
    applications = base.select_related(
        'program_opting', 'status', 'verification_status', 'state', 'district',
        'qualification', 'training_partner'
    ).prefetch_related('documents').order_by(order_by)

    # ---- Result count + pagination ----
    total_count = applications.count()
    filter_params = {
        'search': request.GET.get('search', ''),
        'program': request.GET.get('program', ''),
        'qualification': request.GET.get('qualification', ''),
        'gender': request.GET.get('gender', ''),
        'state': request.GET.get('state', ''),
        'district': request.GET.get('district', ''),
        'institution': request.GET.get('institution', ''),
        'training_partner': request.GET.get('training_partner', ''),
        'app_status': request.GET.get('app_status', ''),
        'verification': request.GET.get('verification', ''),
        'year_of_passing': request.GET.get('year_of_passing', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
    }
    filters_active = any(v for v in filter_params.values())
    paginator = Paginator(applications, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    _attach_photo_urls(page_obj.object_list)

    preserve = dict(filter_params)
    preserve['sort'] = request.GET.get('sort', '')
    query_string = '&'.join(f'{k}={v}' for k, v in preserve.items() if v)

    institutions = (
        StudentApplication.objects.order_by('institution')
        .exclude(institution__isnull=True).exclude(institution='')
        .values_list('institution', flat=True).distinct()
    )[:300]

    context = {
        'title': 'Student Directory',
        'page_obj': page_obj,
        'total_count': total_count,
        'filters_active': filters_active,
        'query_string': query_string,
        'total_students': total_students,
        'submitted_count': submitted_count,
        'draft_count': draft_count,
        'verified_count': verified_count,
        'correction_count': correction_count,
        'pending_verification_count': pending_verification_count,
        'programs': Program.objects.filter(is_active=True),
        'qualifications': Qualification.objects.filter(is_active=True),
        'genders': StudentApplication.GENDER_CHOICES,
        'states': State.objects.filter(is_active=True).order_by('name'),
        'districts': District.objects.filter(is_active=True).order_by('name'),
        'institutions': institutions,
        'training_partners': TrainingPartner.objects.filter(is_active=True),
        'application_statuses': ApplicationStatus.objects.order_by('display_order', 'name'),
        'verification_statuses': VerificationStatus.objects.order_by('display_order', 'name'),
        'passing_years': (
            StudentApplication.objects.exclude(year_of_passing__isnull=True)
            .order_by('-year_of_passing')
            .values_list('year_of_passing', flat=True).distinct()
        ),
        'current_search': request.GET.get('search', ''),
        'current_program': request.GET.get('program', ''),
        'current_qualification': request.GET.get('qualification', ''),
        'current_gender': request.GET.get('gender', ''),
        'current_state': request.GET.get('state', ''),
        'current_district': request.GET.get('district', ''),
        'current_institution': request.GET.get('institution', ''),
        'current_training_partner': request.GET.get('training_partner', ''),
        'current_app_status': request.GET.get('app_status', ''),
        'current_verification': request.GET.get('verification', ''),
        'current_year_of_passing': request.GET.get('year_of_passing', ''),
        'current_date_from': request.GET.get('date_from', ''),
        'current_date_to': request.GET.get('date_to', ''),
        'current_sort': request.GET.get('sort', 'registration_desc'),
        'current_program_name': _name(Program, request.GET.get('program')),
        'current_qualification_name': _name(Qualification, request.GET.get('qualification')),
        'current_state_name': _name(State, request.GET.get('state')),
        'current_district_name': _name(District, request.GET.get('district')),
        'current_training_partner_name': _name(TrainingPartner, request.GET.get('training_partner')),
        'current_app_status_name': _status_name(request.GET.get('app_status')),
        'current_verification_name': _status_name(request.GET.get('verification')),
    }
    return render(request, 'dashboard/student_directory.html', context)


@login_required
@user_passes_test(lambda u: u.is_staff)
def student_profile(request, app_id):
    """
    Complete student profile: personal/contact/education/identification details,
    application & verification status, uploaded documents, and a timeline built
    from the application's tracked timestamps. Read-only.
    """
    from django.shortcuts import get_object_or_404
    app = get_object_or_404(
        StudentApplication.objects.select_related(
            'religion', 'community', 'marital_status', 'ex_serviceman', 'occupation',
            'country', 'state', 'district', 'program_opting', 'qualification',
            'status', 'verification_status', 'training_partner'
        ).prefetch_related('documents'),
        id=app_id,
    )

    documents = app.documents.all()
    doc_groups = {}
    for doc in documents:
        doc_groups.setdefault(doc.doc_type_display, []).append(doc)

    # Timeline events (newest first) built from real tracked timestamps only
    events = []
    for doc in documents:
        events.append({
            'time': doc.uploaded_at,
            'title': f'Document uploaded: {doc.doc_type_display}',
            'detail': doc.original_filename or '',
        })
    events.append({
        'time': app.submission_date,
        'title': 'Application created' if app.is_draft else 'Application submitted',
        'detail': app.application_number or '',
    })
    if app.verification_status:
        events.append({
            'time': app.last_updated,
            'title': f'Verification: {app.verification_status.name}',
            'detail': '',
        })
    if app.approval_date:
        events.append({
            'time': app.approval_date,
            'title': f'Approval: {app.status.name if app.status else "Approved"}',
            'detail': app.approved_by.get_full_name() if app.approved_by else '',
        })
    events.sort(key=lambda e: e['time'] or app.submission_date, reverse=True)

    context = {
        'title': f'Student Profile — {app.full_name or app.application_number}',
        'app': app,
        'doc_groups': doc_groups,
        'events': events,
    }
    return render(request, 'dashboard/student_profile.html', context)


def _advanced_search_filter(request, base):
    """Applies every Advanced Search condition with AND logic."""
    def getval(key):
        return request.GET.get(key, '').strip()

    app_no = getval('application_number')
    name = getval('full_name')
    mobile = getval('mobile_number')
    email = getval('email')
    aadhaar = getval('aadhaar_number')
    gender = getval('gender')
    state_id = getval('state')
    district_id = getval('district')
    area_village = getval('area_village')
    institution = getval('institution')
    qualification_id = getval('qualification')
    program_id = getval('program')
    partner_id = getval('training_partner')
    batch_code = getval('batch_code')
    abc_id = getval('abc_id')
    app_status = getval('application_status')
    dob_from = getval('dob_from')
    dob_to = getval('dob_to')
    age_min = getval('age_min')
    age_max = getval('age_max')
    reg_from = getval('reg_from')
    reg_to = getval('reg_to')
    sub_from = getval('sub_from')
    sub_to = getval('sub_to')

    if app_no:
        base = base.filter(application_number__icontains=app_no)
    if name:
        base = base.filter(full_name__icontains=name)
    if mobile:
        base = base.filter(mobile_number__icontains=mobile)
    if email:
        base = base.filter(email__icontains=email)
    if aadhaar:
        base = base.filter(aadhaar_number__icontains=aadhaar)
    if gender:
        base = base.filter(gender=gender)
    if state_id:
        base = base.filter(state_id=state_id)
    if district_id:
        base = base.filter(district_id=district_id)
    if area_village:
        base = base.filter(area_village_name__icontains=area_village)
    if institution:
        base = base.filter(institution__icontains=institution)
    if qualification_id:
        base = base.filter(qualification_id=qualification_id)
    if program_id:
        base = base.filter(program_opting_id=program_id)
    if partner_id:
        base = base.filter(training_partner_id=partner_id)
    if batch_code:
        base = base.filter(batch_code__icontains=batch_code)
    if abc_id:
        base = base.filter(abc_id__icontains=abc_id)
    if app_status == 'draft':
        base = base.filter(status__code='INCOMPLETE')
    elif app_status == 'submitted':
        base = base.filter(status__code__in=SUBMITTED_STATUS_CODES)

    today = timezone.localdate()
    if dob_from:
        try:
            base = base.filter(dob__gte=timezone.datetime.strptime(dob_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if dob_to:
        try:
            base = base.filter(dob__lte=timezone.datetime.strptime(dob_to, '%Y-%m-%d').date())
        except ValueError:
            pass
    if age_min:
        try:
            cutoff = today.replace(year=today.year - int(age_min))
            base = base.filter(dob__lte=cutoff)
        except (ValueError, TypeError):
            pass
    if age_max:
        try:
            cutoff = today.replace(year=today.year - int(age_max))
            base = base.filter(dob__gte=cutoff)
        except (ValueError, TypeError):
            pass

    def apply_date(qs, frm, to, field):
        if frm:
            try:
                qs = qs.filter(**{f'{field}__date__gte': timezone.datetime.strptime(frm, '%Y-%m-%d').date()})
            except ValueError:
                pass
        if to:
            try:
                qs = qs.filter(**{f'{field}__date__lte': timezone.datetime.strptime(to, '%Y-%m-%d').date()})
            except ValueError:
                pass
        return qs

    base = apply_date(base, reg_from, reg_to, 'submission_date')
    base = apply_date(base, sub_from, sub_to, 'submission_date')
    return base


@login_required
@user_passes_test(lambda u: u.is_staff)
def student_advanced_search(request):
    """
    Advanced Search: multi-condition search across the student database with
    AND-combined filters, saved searches (session), export and bulk selection.
    """
    from django.http import JsonResponse

    # ---- Save / delete saved searches (session-based, per staff user) ----
    saved_key = 'saved_student_searches'
    saved = request.session.get(saved_key, [])
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        action = request.POST.get('action')
        if action == 'save':
            name = request.POST.get('name', '').strip()[:100]
            qs = request.POST.get('query_string', '')
            if name and qs:
                saved.append({'name': name, 'query_string': qs})
                request.session[saved_key] = saved
                return JsonResponse({'ok': True, 'count': len(saved)})
            return JsonResponse({'ok': False, 'error': 'Name and query string required.'}, status=400)
        if action == 'delete':
            idx = request.POST.get('index')
            try:
                idx = int(idx)
                if 0 <= idx < len(saved):
                    saved.pop(idx)
                    request.session[saved_key] = saved
                    return JsonResponse({'ok': True, 'count': len(saved)})
            except (TypeError, ValueError):
                pass
            return JsonResponse({'ok': False, 'error': 'Invalid index.'}, status=400)
        return JsonResponse({'ok': False, 'error': 'Unknown action.'}, status=400)

    base = _advanced_search_filter(request, StudentApplication.objects.all())
    applications = base.select_related(
        'program_opting', 'status', 'verification_status', 'state', 'district',
        'qualification', 'training_partner'
    ).prefetch_related('documents').order_by('-submission_date')

    total_count = applications.count()
    paginator = Paginator(applications, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    _attach_photo_urls(page_obj.object_list)

    current = {k: request.GET.get(k, '') for k in [
        'application_number', 'full_name', 'mobile_number', 'email', 'aadhaar_number',
        'gender', 'state', 'district', 'area_village', 'institution', 'qualification',
        'program', 'training_partner', 'batch_code', 'abc_id', 'application_status',
        'dob_from', 'dob_to', 'age_min', 'age_max', 'reg_from', 'reg_to', 'sub_from', 'sub_to',
    ]}
    filters_active = any(v for v in current.values())
    query_string = '&'.join(f'{k}={v}' for k, v in current.items() if v)

    context = {
        'title': 'Advanced Search',
        'page_obj': page_obj,
        'total_count': total_count,
        'filters_active': filters_active,
        'query_string': query_string,
        'saved_searches': saved,
        'statuses': [
            {'value': 'draft', 'name': 'Draft'},
            {'value': 'submitted', 'name': 'Submitted'},
        ],
        'genders': StudentApplication.GENDER_CHOICES,
        'states': State.objects.filter(is_active=True).order_by('name'),
        'districts': District.objects.filter(is_active=True).order_by('name'),
        'qualifications': Qualification.objects.filter(is_active=True),
        'programs': Program.objects.filter(is_active=True),
        'training_partners': TrainingPartner.objects.filter(is_active=True),
        'current': current,
        'current_program_name': _name(Program, current['program']),
        'current_qualification_name': _name(Qualification, current['qualification']),
        'current_state_name': _name(State, current['state']),
        'current_district_name': _name(District, current['district']),
        'current_training_partner_name': _name(TrainingPartner, current['training_partner']),
    }
    return render(request, 'dashboard/student_advanced_search.html', context)


@login_required
@user_passes_test(lambda u: u.is_staff)
def advanced_search_export(request):
    """Exports the current Advanced Search result set (CSV or Excel)."""
    from django.http import HttpResponse
    from utilities.export_generator import generate_full_csv_export, generate_full_excel_export

    applications = _advanced_search_filter(request, StudentApplication.objects.all())
    fmt = request.GET.get('format', 'csv').lower()
    if fmt == 'excel':
        data = generate_full_excel_export(applications, request)
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = 'advanced_search_students.xlsx'
    else:
        data = generate_full_csv_export(applications, request)
        content_type = 'text/csv; charset=utf-8'
        filename = 'advanced_search_students.csv'
    response = HttpResponse(data, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@user_passes_test(lambda u: u.is_staff)
def student_bulk_operations(request):
    """
    Bulk Operations: select multiple students and run operations on ONLY the
    selected IDs. Includes export, document ZIPs, change status, assign
    training partner / batch code, request correction, email notification.
    """
    base = StudentApplication.objects.all()

    search_query = request.GET.get('search', '').strip()
    if search_query:
        base = base.filter(
            Q(application_number__icontains=search_query) |
            Q(full_name__icontains=search_query) |
            Q(mobile_number__icontains=search_query) |
            Q(aadhaar_number__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    status = request.GET.get('status')
    if status == 'draft':
        base = base.filter(status__code='INCOMPLETE')
    elif status == 'submitted':
        base = base.filter(status__code__in=SUBMITTED_STATUS_CODES)

    applications = base.select_related('program_opting', 'status', 'verification_status').order_by('-submission_date')
    total_count = applications.count()
    paginator = Paginator(applications, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'title': 'Bulk Operations',
        'page_obj': page_obj,
        'total_count': total_count,
        'current_search': request.GET.get('search', ''),
        'current_status': request.GET.get('status', ''),
        'statuses': [
            {'value': 'draft', 'name': 'Draft'},
            {'value': 'submitted', 'name': 'Submitted'},
        ],
        'application_statuses': ApplicationStatus.objects.exclude(code__in=['APPROVED', 'REJECTED']).order_by('name'),
        'training_partners': TrainingPartner.objects.filter(is_active=True),
        'batch_codes': BatchCode.objects.select_related('training_partner').filter(is_active=True).order_by('code'),
    }
    return render(request, 'dashboard/student_bulk_operations.html', context)


@login_required
@user_passes_test(lambda u: u.is_staff)
def bulk_change_status(request):
    """Changes the application status of ONLY selected applications."""
    from django.db import transaction
    from django.http import JsonResponse

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed.'}, status=405)
    code = request.POST.get('status', '').strip()
    if not code:
        return JsonResponse({'error': 'Choose a status first.'}, status=400)
    target = ApplicationStatus.objects.filter(code=code).first()
    if not target:
        return JsonResponse({'error': 'Invalid application status.'}, status=400)
    ids = _collect_selected_ids(request)
    if not ids:
        return JsonResponse({'error': 'Please select at least one student.'}, status=400)
    queryset = _selected_applications_queryset(request)
    total = queryset.count()
    with transaction.atomic():
        processed = queryset.update(status=target)
    skipped = total - processed
    if processed:
        _write_audit_log(request, 'BULK_CHANGE_STATUS', f'Changed {processed} application(s) to {target.name}.', ids)
    return JsonResponse({'processed': processed, 'skipped': skipped, 'status': target.name})


@login_required
@user_passes_test(lambda u: u.is_staff)
def bulk_assign_training_partner(request):
    """Assigns a Training Partner/ATC to ONLY selected applications."""
    from django.db import transaction
    from django.http import JsonResponse

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed.'}, status=405)
    partner_id = request.POST.get('training_partner', '').strip()
    if not partner_id:
        return JsonResponse({'error': 'Select a training partner.'}, status=400)
    partner = TrainingPartner.objects.filter(id=partner_id).first()
    if not partner:
        return JsonResponse({'error': 'Invalid training partner.'}, status=400)
    ids = _collect_selected_ids(request)
    if not ids:
        return JsonResponse({'error': 'Please select at least one student.'}, status=400)
    queryset = _selected_applications_queryset(request)
    total = queryset.count()
    with transaction.atomic():
        processed = queryset.update(training_partner=partner)
    skipped = total - processed
    if processed:
        _write_audit_log(request, 'BULK_ASSIGN_PARTNER', f'Assigned {processed} application(s) to {partner.name}.', ids)
    return JsonResponse({'processed': processed, 'skipped': skipped, 'training_partner': partner.name})


@login_required
@user_passes_test(lambda u: u.is_staff)
def bulk_assign_batch_code(request):
    """Assigns a batch code to ONLY selected applications."""
    from django.db import transaction
    from django.http import JsonResponse

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed.'}, status=405)
    batch_code = request.POST.get('batch_code', '').strip()[:50]
    if not batch_code:
        return JsonResponse({'error': 'Enter a batch code.'}, status=400)
    ids = _collect_selected_ids(request)
    if not ids:
        return JsonResponse({'error': 'Please select at least one student.'}, status=400)
    queryset = _selected_applications_queryset(request)
    total = queryset.count()
    with transaction.atomic():
        processed = queryset.update(batch_code=batch_code)
    skipped = total - processed
    if processed:
        _write_audit_log(request, 'BULK_ASSIGN_BATCH', f'Assigned batch {batch_code} to {processed} application(s).', ids)
    return JsonResponse({'processed': processed, 'skipped': skipped, 'batch_code': batch_code})


@login_required
@user_passes_test(lambda u: u.is_staff)
def bulk_download_applications(request):
    """
    Generates a CSC_Selected_Applications_*.zip of the acknowledgement PDFs for
    ONLY the selected applications. Students without an acknowledgement PDF are skipped.
    """
    import zipfile
    import io
    import re
    from django.http import HttpResponse, JsonResponse

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed.'}, status=405)

    def safe_name(value):
        name = re.sub(r'[^\w\- ]', '_', str(value or ''))
        name = re.sub(r'[\s]+', '_', name)
        name = name.strip('._')
        return name or 'Application'

    ids = _collect_selected_ids(request)
    if not ids:
        return JsonResponse({'error': 'Please select at least one student.'}, status=400)

    queryset = _selected_applications_queryset(request)
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for app in queryset:
            if not app.acknowledgement_pdf or not app.acknowledgement_pdf.name:
                continue
            try:
                if os.path.exists(app.acknowledgement_pdf.path):
                    folder = f"CSC_Selected_Applications/{safe_name(app.application_number) or f'App_{app.id}'}/"
                    zip_file.write(app.acknowledgement_pdf.path, arcname=folder + "Acknowledgement.pdf")
            except Exception:
                continue
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="CSC_Selected_Applications_{_bulk_timestamp()}.zip"'
    return response


@login_required
@user_passes_test(lambda u: u.is_staff)
def bulk_notify(request):
    """Emails ONLY selected students who have an email address (fail_silently)."""
    from django.http import JsonResponse
    from django.core.mail import send_mail

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed.'}, status=405)
    subject = request.POST.get('subject', '').strip()[:200]
    message = request.POST.get('message', '').strip()
    if not subject or not message:
        return JsonResponse({'error': 'Subject and message are required.'}, status=400)

    ids = _collect_selected_ids(request)
    if not ids:
        return JsonResponse({'error': 'Please select at least one student.'}, status=400)

    all_selected = _selected_applications_queryset(request)
    total_selected = all_selected.count()
    with_email = all_selected.exclude(email__isnull=True).exclude(email='')
    no_email = total_selected - with_email.count()
    sent = 0
    for app in with_email:
        try:
            send_mail(
                subject,
                f"Dear {app.full_name},\n\n{message}\n\nRegards,\nCSC Admin Team",
                settings.DEFAULT_FROM_EMAIL,
                [app.email],
                fail_silently=True,
            )
            sent += 1
        except Exception:
            continue
    if sent:
        _write_audit_log(request, 'BULK_NOTIFY', f'Sent {sent} email(s) to selected students.', ids)
    return JsonResponse({'sent': sent, 'skipped': with_email.count() - sent, 'no_email': no_email})


@login_required
@user_passes_test(lambda u: u.is_staff)
def bulk_print(request):
    """
    Printable records page for ONLY the selected application IDs (GET ?ids=1,2,3).
    Renders a print-friendly template and triggers window.print().
    """
    ids = request.GET.get('ids', '')
    id_list = []
    for raw in ids.split(','):
        try:
            pk = int(raw.strip())
        except (TypeError, ValueError):
            continue
        if pk > 0:
            id_list.append(pk)
    applications = StudentApplication.objects.filter(id__in=id_list).select_related(
        'program_opting', 'status', 'verification_status', 'state', 'district', 'qualification'
    ).order_by('-submission_date')
    context = {'title': 'Print Selected Students', 'applications': applications, 'now': timezone.now()}
    return render(request, 'dashboard/student_print.html', context)


@login_required
@user_passes_test(lambda u: u.is_staff)
def verification_desk(request):
    return render(request, 'dashboard/placeholder.html', {'title': 'Verification Desk'})

@login_required
@user_passes_test(lambda u: u.is_staff)
def reports_dashboard(request):
    from django.shortcuts import redirect
    return redirect('report_daily')

@login_required
@user_passes_test(lambda u: u.is_staff)
def export_center(request):
    """
    Export Center: centralizes every export/download operation for the portal.
    Tracks each generated export in ExportJob for history + audit trail.
    """
    from dashboard.models import ExportJob
    from django.core.paginator import Paginator as _Paginator

    jobs_qs = ExportJob.objects.select_related('requested_by')

    # Export history search (id / type / format / user / status)
    history_search = request.GET.get('history_search', '').strip()
    if history_search:
        jobs_qs = jobs_qs.filter(
            Q(export_type__icontains=history_search) |
            Q(file_format__icontains=history_search) |
            Q(status__icontains=history_search) |
            Q(requested_by__username__icontains=history_search)
        )

    today = timezone.localdate()
    stats = {
        'total': jobs_qs.count(),
        'today': jobs_qs.filter(created_at__date=today).count(),
        'completed': jobs_qs.filter(status='completed').count(),
        'failed': jobs_qs.filter(status='failed').count(),
        'processing': jobs_qs.filter(status='processing').count(),
    }

    history_paginator = _Paginator(jobs_qs, 20)
    history_page = history_paginator.get_page(request.GET.get('history_page'))

    # Application export filter dropdown data
    filters = _export_center_filters(request)

    context = {
        'title': 'Export Center',
        'stats': stats,
        'history_page': history_page,
        'history_search': history_search,
        'f': filters,
        'filters_active': any(v for v in filters.values()),
        'states': State.objects.filter(is_active=True).order_by('name'),
        'districts': District.objects.filter(is_active=True).order_by('name'),
        'programs': Program.objects.filter(is_active=True).order_by('name'),
        'qualifications': Qualification.objects.filter(is_active=True).order_by('name'),
        'training_partners': TrainingPartner.objects.filter(is_active=True).order_by('name'),
        'batch_codes': BatchCode.objects.select_related('training_partner').filter(is_active=True).order_by('code'),
        'document_types': _DOCUMENT_TYPE_CHOICES,
        'master_data_models': _MASTER_DATA_MODELS,
        'export_job': None,
    }
    return render(request, 'dashboard/export_center.html', context)


def _export_center_filters(request):
    """Collects Export Center application-filter values from POST (generate) or GET (page)."""
    source = request.POST if request.method == 'POST' else request.GET
    return {
        'search': source.get('search', ''),
        'application_number': source.get('application_number', ''),
        'name': source.get('name', ''),
        'mobile': source.get('mobile', ''),
        'email': source.get('email', ''),
        'aadhaar': source.get('aadhaar', ''),
        'status': source.get('status', ''),
        'program': source.get('program', ''),
        'qualification': source.get('qualification', ''),
        'training_partner': source.get('training_partner', ''),
        'batch_code': source.get('batch_code', ''),
        'state': source.get('state', ''),
        'district': source.get('district', ''),
        'date_from': source.get('date_from', ''),
        'date_to': source.get('date_to', ''),
    }


def _apply_export_filters(base, f):
    """Applies Export Center application filters (read-only) to a queryset."""
    qs = base
    if f.get('search'):
        q = Q(application_number__icontains=f['search']) | \
            Q(full_name__icontains=f['search']) | \
            Q(mobile_number__icontains=f['search']) | \
            Q(email__icontains=f['search']) | \
            Q(aadhaar_number__icontains=f['search'])
        qs = qs.filter(q)
    if f.get('application_number'):
        qs = qs.filter(application_number__icontains=f['application_number'])
    if f.get('name'):
        qs = qs.filter(full_name__icontains=f['name'])
    if f.get('mobile'):
        qs = qs.filter(mobile_number__icontains=f['mobile'])
    if f.get('email'):
        qs = qs.filter(email__icontains=f['email'])
    if f.get('aadhaar'):
        qs = qs.filter(aadhaar_number__icontains=f['aadhaar'])
    if f.get('status') == 'draft':
        qs = qs.filter(status__code='INCOMPLETE')
    elif f.get('status') == 'submitted':
        qs = qs.filter(status__code__in=SUBMITTED_STATUS_CODES)
    elif f.get('status') == 'pending':
        qs = qs.filter(status__code='PENDING')
    if f.get('program'):
        qs = qs.filter(program_opting_id=f['program'])
    if f.get('qualification'):
        qs = qs.filter(qualification_id=f['qualification'])
    if f.get('training_partner'):
        qs = qs.filter(training_partner_id=f['training_partner'])
    if f.get('batch_code'):
        qs = qs.filter(batch_code__icontains=f['batch_code'])
    if f.get('state'):
        qs = qs.filter(state_id=f['state'])
    if f.get('district'):
        qs = qs.filter(district_id=f['district'])
    if f.get('date_from'):
        try:
            qs = qs.filter(submission_date__date__gte=timezone.datetime.strptime(f['date_from'], '%Y-%m-%d').date())
        except ValueError:
            pass
    if f.get('date_to'):
        try:
            qs = qs.filter(submission_date__date__lte=timezone.datetime.strptime(f['date_to'], '%Y-%m-%d').date())
        except ValueError:
            pass
    return qs


_DOCUMENT_TYPE_CHOICES = [
    ('Passport_Photo', 'Photograph'),
    ('Signature', 'Signature'),
    ('Left_Thumb', 'Thumb Impression'),
    ('Aadhaar', 'Aadhaar'),
    ('ABC_ID', 'ABC ID'),
    ('Father_Signature', "Father's/Guardian's Signature"),
    ('Community_Certificate', 'Community Certificate'),
    ('Additional_Documents', 'Additional Documents'),
    ('Registration_Screenshot', 'Registration Screenshot'),
    ('Supporting_Documents', 'Supporting Documents'),
]

# Maps the many stored doc_type spellings to one canonical folder name.
_DOCUMENT_FOLDER_MAP = {
    'passport_photo': 'Passport_Photo',
    'passport': 'Passport_Photo',
    'signature': 'Signature',
    'thumb_impression': 'Left_Thumb',
    'left_thumb': 'Left_Thumb',
    'thumb': 'Left_Thumb',
    'aadhaar': 'Aadhaar',
    'aadhaar_pdf': 'Aadhaar',
    'abc': 'ABC_ID',
    'abc_id': 'ABC_ID',
    'abc_screenshot': 'ABC_ID',
    'father_signature': 'Father_Signature',
    'community': 'Community_Certificate',
    'community_certificate': 'Community_Certificate',
    'additional_documents': 'Additional_Documents',
    'registration': 'Registration_Screenshot',
    'registration_screenshot': 'Registration_Screenshot',
    'supporting': 'Supporting_Documents',
    'supporting_documents': 'Supporting_Documents',
}


def _doc_folder(doc_type):
    """Resolves any stored doc_type spelling to its canonical folder name."""
    key = str(doc_type or '').strip().lower()
    return _DOCUMENT_FOLDER_MAP.get(key)


def _build_applications_export(fmt, f, request):
    """Builds application export bytes + content-type for csv/xlsx/pdf/zip."""
    import io as _io
    import csv as _csv
    from utilities.export_generator import extract_application_row, EXPORT_COLUMNS

    queryset = _apply_export_filters(StudentApplication.objects.all(), f)
    count = queryset.count()

    if fmt == 'zip':
        return _build_documents_zip(queryset, f.get('document_types'), f.get('include_documents'))

    if fmt == 'csv':
        from utilities.export_generator import generate_full_csv_export
        data = generate_full_csv_export(queryset, request)
        return data, 'text/csv; charset=utf-8', 'csv', count

    if fmt == 'xlsx':
        from utilities.export_generator import generate_full_excel_export
        data = generate_full_excel_export(queryset, request)
        return data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'xlsx', count

    if fmt == 'pdf':
        # Compact column subset so the PDF stays readable in landscape.
        headers = [
            'Application No', 'Name', 'Father Name', 'Mobile', 'Email', 'Aadhaar',
            'DOB', 'Gender', 'State', 'District', 'Program', 'Qualification',
            'Training Partner', 'Batch Code', 'Status', 'Submission Date',
        ]
        rows = []
        for app in queryset.select_related(
            'state', 'district', 'program_opting', 'qualification',
            'training_partner', 'status', 'verification_status'
        ):
            rows.append([
                app.application_number or '', app.full_name or '', app.father_name or '',
                app.mobile_number or '', app.email or '', app.aadhaar_number or '',
                app.dob.strftime('%d-%m-%Y') if app.dob else '', app.gender or '',
                (app.state.name if app.state else ''), (app.district.name if app.district else ''),
                (app.program_opting.name if app.program_opting else ''),
                (app.qualification.name if app.qualification else ''),
                (app.training_partner.name if app.training_partner else ''),
                app.batch_code or '', (app.status.name if app.status else ''),
                app.submission_date.strftime('%d-%m-%Y %H:%M') if app.submission_date else '',
            ])
        from django.template.loader import render_to_string
        html = render_to_string('dashboard/report_export_pdf.html', {
            'title': 'Student Applications Export',
            'headers': headers,
            'rows': rows,
            'generated_at': timezone.now(),
        })
        buffer = _io.BytesIO()
        from xhtml2pdf import pisa
        pisa.CreatePDF(html, dest=buffer, encoding='utf-8')
        return buffer.getvalue(), 'application/pdf', 'pdf', count

    raise ValueError('Unsupported application export format.')


def _build_documents_zip(queryset, doc_types, include_documents):
    """Builds a ZIP of selected documents for the filtered applications."""
    import io as _io
    import zipfile
    import os as _os
    import re as _re

    buffer = _io.BytesIO()
    if not include_documents:
        # No documents requested -> build an empty-but-valid ZIP with a notice.
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr('NO_DOCUMENTS_REQUESTED.txt', 'No documents were requested for this export.')
        return buffer.getvalue(), 'application/zip', 'zip', 0

    def safe_name(value):
        name = _re.sub(r'[^\w\- ]', '_', str(value or ''))
        name = _re.sub(r'[\s]+', '_', name)
        return name.strip('._') or 'Application'

    queryset = queryset.prefetch_related('documents')
    if doc_types:
        selected = set()
        for dt in doc_types:
            canonical = _doc_folder(dt) or str(dt).strip()
            selected.add(canonical)
    else:
        selected = set(dt[0] for dt in _DOCUMENT_TYPE_CHOICES)
    count = 0
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for app in queryset:
            app_no = safe_name(app.application_number) if app.application_number else f"App_{app.id}"
            folder = f"CSC_Selected_Documents/{app_no}/"
            for doc in app.documents.all():
                folder_name = _doc_folder(doc.doc_type) or safe_name(doc.doc_type)
                if folder_name not in selected:
                    continue
                for field in ('original_file', 'compressed_file'):
                    file_field = getattr(doc, field, None)
                    if file_field and file_field.name:
                        try:
                            if _os.path.exists(file_field.path):
                                ext = _os.path.splitext(file_field.name)[1] or ''
                                suffix = 'optimized' if field == 'compressed_file' else 'original'
                                zf.write(file_field.path, arcname=f"{folder}{folder_name}/{suffix}{ext}")
                                count += 1
                        except Exception:
                            continue
    buffer.seek(0)
    return buffer.getvalue(), 'application/zip', 'zip', count


_MASTER_DATA_MODELS = [
    {'key': 'qualifications', 'label': 'Qualifications', 'model': 'Qualification'},
    {'key': 'programs', 'label': 'Programs', 'model': 'Program'},
    {'key': 'training_partners', 'label': 'Training Partners', 'model': 'TrainingPartner'},
    {'key': 'batch_codes', 'label': 'Batch Codes', 'model': 'BatchCode'},
    {'key': 'states', 'label': 'States', 'model': 'State'},
    {'key': 'districts', 'label': 'Districts', 'model': 'District'},
    {'key': 'communities', 'label': 'Communities', 'model': 'Community'},
    {'key': 'occupations', 'label': 'Occupations', 'model': 'Occupation'},
    {'key': 'application_statuses', 'label': 'Application Statuses', 'model': 'ApplicationStatus'},
    {'key': 'institutions', 'label': 'Institutions', 'model': None},
]


def _master_data_queryset(key):
    """Resolves a master-data export key to a queryset of name/code rows."""
    from registrations.models import StudentApplication as _SA
    model_map = {
        'qualifications': Qualification,
        'programs': Program,
        'training_partners': TrainingPartner,
        'batch_codes': BatchCode,
        'states': State,
        'districts': District,
        'communities': Community,
        'occupations': Occupation,
        'application_statuses': ApplicationStatus,
    }
    model = model_map.get(key)
    if model is None and key == 'institutions':
        values = (_SA.objects.exclude(institution__isnull=True)
                  .exclude(institution='')
                  .values_list('institution', flat=True)
                  .distinct().order_by('institution'))
        return list(values)
    if model is None:
        return []
    return list(model.objects.filter(is_active=True).order_by('name'))


def _build_master_data_export(key, fmt):
    """Builds master-data export bytes for a given key + format."""
    import io as _io
    import csv as _csv

    data = _master_data_queryset(key)
    if isinstance(data, list) and data and isinstance(data[0], str):
        headers = ['Institution']
        rows = [[name] for name in data]
    else:
        headers = ['Name', 'Code']
        rows = [[(getattr(x, 'name', '') or ''), (getattr(x, 'code', '') or '')] for x in data]

    count = len(rows)

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Master Data"
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
        for row in rows:
            ws.append(row)
        buffer = _io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'xlsx', count

    buffer = _io.StringIO()
    writer = _csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    return buffer.getvalue().encode('utf-8'), 'text/csv; charset=utf-8', 'csv', count


def _build_audit_logs_export(fmt):
    """Builds audit-log export bytes."""
    import io as _io
    import csv as _csv
    from dashboard.models import AuditLog

    headers = ['ID', 'User', 'Action', 'Detail', 'Application IDs', 'Date/Time']
    rows = []
    for log in AuditLog.objects.select_related('user').order_by('-created_at'):
        rows.append([
            log.id, (log.user.username if log.user else 'anonymous'), log.action,
            log.detail, log.application_ids,
            log.created_at.strftime('%d-%m-%Y %H:%M:%S') if log.created_at else '',
        ])

    count = len(rows)
    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Logs"
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
        for row in rows:
            ws.append(row)
        buffer = _io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'xlsx', count

    buffer = _io.StringIO()
    writer = _csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    return buffer.getvalue().encode('utf-8'), 'text/csv; charset=utf-8', 'csv', count


@login_required
@user_passes_test(lambda u: u.is_staff)
def export_center_generate(request):
    """
    POST handler that runs an export (applications/students/documents/master_data/
    audit_logs) using live data, records an ExportJob, saves the generated file,
    and redirects to the download URL so the admin gets the file immediately.
    """
    from dashboard.models import ExportJob
    from django.core.files.base import ContentFile
    from django.core.files.storage import default_storage
    from django.shortcuts import redirect
    from django.urls import reverse

    if request.method != 'POST':
        return redirect('export_center')

    export_type = request.POST.get('export_type', 'applications')
    fmt = request.POST.get('format', 'csv')

    perm_map = {
        'applications': 'dashboard.export_applications',
        'students': 'dashboard.export_student_data',
        'documents': 'dashboard.export_documents',
        'reports': 'dashboard.export_applications',
        'master_data': 'dashboard.export_master_data',
        'audit_logs': 'dashboard.export_audit_logs',
    }
    perm = perm_map.get(export_type)
    if perm and not (request.user.is_superuser or request.user.has_perm(perm)):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("You do not have permission to run this export.")

    if fmt not in ('csv', 'xlsx', 'pdf', 'zip'):
        fmt = 'csv'

    f = _export_center_filters(request)

    job = ExportJob.objects.create(
        export_type=export_type,
        file_format=fmt,
        requested_by=request.user if request.user.is_authenticated else None,
        filters={k: v for k, v in f.items() if v},
        status='processing',
    )

    try:
        if export_type == 'documents':
            doc_types = request.POST.getlist('document_types')
            f['document_types'] = doc_types
            f['include_documents'] = True
            data, content_type, ext, count = _build_applications_export('zip', f, request)
            filename = f"CSC_Selected_Documents_{_bulk_timestamp()}.zip"
        elif export_type == 'master_data':
            key = request.POST.get('master_data_key', 'qualifications')
            data, content_type, ext, count = _build_master_data_export(key, fmt)
            label = next((m['label'] for m in _MASTER_DATA_MODELS if m['key'] == key), key)
            filename = f"{label.replace(' ', '_')}_{_bulk_timestamp()}.{ext}"
        elif export_type == 'audit_logs':
            data, content_type, ext, count = _build_audit_logs_export(fmt)
            filename = f"CSC_Audit_Logs_{_bulk_timestamp()}.{ext}"
        elif export_type == 'students':
            data, content_type, ext, count = _build_students_export(fmt, f, request)
            filename = f"CSC_Selected_Students_{_bulk_timestamp()}.{ext}"
        elif export_type == 'reports':
            data, content_type, ext, count = _build_reports_export(fmt, f, request)
            filename = f"CSC_Report_{_bulk_timestamp()}.{ext}"
        else:
            f['include_documents'] = request.POST.get('include_documents') == 'on'
            f['document_types'] = request.POST.getlist('document_types')
            data, content_type, ext, count = _build_applications_export(fmt, f, request)
            filename = f"CSC_Selected_Students_{_bulk_timestamp()}.{ext}"

        job.record_count = count
        saved_name = default_storage.save(f'exports/{filename}', ContentFile(data))
        job.file.name = saved_name
        job.status = 'completed'
        job.completed_at = timezone.now()
        job.expires_at = timezone.now() + timedelta(days=7)
        job.save()

        _write_audit_log(
            request,
            'EXPORT_CENTER',
            f'Exported {job.export_type} as {job.file_format} ({count} record(s)) -> {job.display_id}.',
        )
        return redirect(reverse('export_center_download', args=[job.pk]))
    except Exception as exc:
        job.status = 'failed'
        job.error_message = str(exc)[:2000]
        job.completed_at = timezone.now()
        job.save()
        _write_audit_log(
            request,
            'EXPORT_FAILED',
            f'Export {job.export_type}/{job.file_format} failed: {str(exc)[:300]}',
        )
        from django.contrib import messages
        messages.error(request, f"Export failed: {exc}")
        return redirect('export_center')


def _build_students_export(fmt, f, request):
    """Builds a Student Directory export (compact column set)."""
    import io as _io
    import csv as _csv
    from django.template.loader import render_to_string

    queryset = _apply_export_filters(StudentApplication.objects.all(), f)
    count = queryset.count()

    headers = [
        'Application No', 'Name', "Father's Name", "Mother's Name", 'Date of Birth',
        'Gender', 'Mobile', 'Alternative Mobile', 'Email', 'Aadhaar',
        'Address', 'State', 'District', 'Pincode', 'Program', 'Qualification',
        'Institution', 'Training Partner', 'Batch Code', 'Status',
        'Verification Status', 'Submission Date',
    ]
    apps = queryset.select_related(
        'state', 'district', 'program_opting', 'qualification',
        'training_partner', 'status', 'verification_status'
    )
    rows = []
    for app in apps:
        rows.append([
            app.application_number or '', app.full_name or '', app.father_name or '',
            app.mother_name or '', app.dob.strftime('%d-%m-%Y') if app.dob else '',
            app.gender or '', app.mobile_number or '', app.alternative_mobile or '',
            app.email or '', app.aadhaar_number or '', app.communication_address or '',
            (app.state.name if app.state else ''), (app.district.name if app.district else ''),
            app.pincode or '', (app.program_opting.name if app.program_opting else ''),
            (app.qualification.name if app.qualification else ''), app.institution or '',
            (app.training_partner.name if app.training_partner else ''), app.batch_code or '',
            (app.status.name if app.status else ''),
            (app.verification_status.name if app.verification_status else ''),
            app.submission_date.strftime('%d-%m-%Y %H:%M') if app.submission_date else '',
        ])

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Directory"
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
        for row in rows:
            ws.append(row)
        buffer = _io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'xlsx', count

    if fmt == 'pdf':
        html = render_to_string('dashboard/report_export_pdf.html', {
            'title': 'Student Directory Export',
            'headers': headers,
            'rows': rows,
            'generated_at': timezone.now(),
        })
        buffer = _io.BytesIO()
        from xhtml2pdf import pisa
        pisa.CreatePDF(html, dest=buffer, encoding='utf-8')
        return buffer.getvalue(), 'application/pdf', 'pdf', count

    buffer = _io.StringIO()
    writer = _csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    return buffer.getvalue().encode('utf-8'), 'text/csv; charset=utf-8', 'csv', count


def _build_reports_export(fmt, f, request):
    """Builds an Application Reports export (summary grouped by day)."""
    import io as _io
    from django.db.models.functions import TruncDate

    qs = _apply_export_filters(StudentApplication.objects.all(), f)
    daily = (
        qs.annotate(day=TruncDate('submission_date'))
        .values('day')
        .annotate(
            total=Count('id'),
            submitted=Count('id', filter=Q(status__code__in=SUBMITTED_STATUS_CODES)),
            draft=Count('id', filter=Q(status__code='INCOMPLETE')),
            pending=Count('id', filter=Q(status__code='PENDING')),
        )
        .order_by('-day')
    )
    headers = ['Date', 'Total', 'Submitted', 'Draft', 'Pending Review']
    rows = [[
        (r['day'].strftime('%d-%m-%Y') if r['day'] else '-'),
        r['total'], r['submitted'], r['draft'], r['pending'],
    ] for r in daily]
    count = len(rows)

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Application Reports"
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
        for row in rows:
            ws.append(row)
        buffer = _io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'xlsx', count

    if fmt == 'pdf':
        from django.template.loader import render_to_string
        html = render_to_string('dashboard/report_export_pdf.html', {
            'title': 'Application Reports Export',
            'headers': headers,
            'rows': rows,
            'generated_at': timezone.now(),
        })
        buffer = _io.BytesIO()
        from xhtml2pdf import pisa
        pisa.CreatePDF(html, dest=buffer, encoding='utf-8')
        return buffer.getvalue(), 'application/pdf', 'pdf', count

    import csv as _csv
    buffer = _io.StringIO()
    writer = _csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    return buffer.getvalue().encode('utf-8'), 'text/csv; charset=utf-8', 'csv', count


@login_required
@user_passes_test(lambda u: u.is_staff)
def export_center_download(request, job_id):
    """
    Serves the generated file for a completed ExportJob. Denies if the job is
    still processing, failed, or expired. Requires download permission.
    """
    from dashboard.models import ExportJob
    from django.http import FileResponse, Http404

    if not (request.user.is_superuser or request.user.has_perm('dashboard.download_export_history')):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("You do not have permission to download exports.")

    job = ExportJob.objects.filter(pk=job_id).first()
    if not job or not job.file:
        raise Http404("Export job not found.")

    if job.status != 'completed':
        return HttpResponse(f"Export is not ready (status: {job.status}).", status=409)
    if job.is_expired:
        return HttpResponse("This export download has expired.", status=410)
    if not os.path.exists(job.file.path):
        raise Http404("Export file missing on disk.")

    response = FileResponse(open(job.file.path, 'rb'))
    response['Content-Disposition'] = f'attachment; filename="{job.filename}"'
    return response

@login_required
@user_passes_test(lambda u: u.is_staff)
def master_data(request):
    return render(request, 'dashboard/placeholder.html', {'title': 'Master Data'})

# ==========================================
# SYSTEM ADMINISTRATION - PLATFORM SETTINGS
# ==========================================

SETTING_DEFINITIONS = [
    # (key, label, category, field_type, placeholder/help, default)
    # Application Settings
    ('application_prefix', 'Application Number Prefix', 'application', 'text', 'Prefix for generated application numbers', 'NIELIT'),
    ('submission_rules', 'Submission Rules', 'application', 'textarea', 'Rules shown to applicants before submission', 'All fields marked * are mandatory.'),
    ('allowed_document_formats', 'Allowed Document Formats', 'application', 'text', 'Comma-separated list', 'jpg, jpeg, png, pdf'),
    # Email Settings
    ('smtp_host', 'SMTP Host', 'email', 'text', 'e.g. smtp.gmail.com', ''),
    ('smtp_port', 'SMTP Port', 'email', 'number', 'e.g. 587', '587'),
    ('smtp_user', 'SMTP Username', 'email', 'text', 'Account used to send mail', ''),
    ('smtp_password', 'SMTP Password', 'email', 'password', 'App password / SMTP secret', ''),
    ('sender_name', 'Sender Name', 'email', 'text', 'Display name on outbound emails', ''),
    ('sender_email', 'Sender Email', 'email', 'email', 'From-address for outbound emails', ''),
    ('email_templates', 'Email Templates', 'email', 'textarea', 'Default template bodies', 'Welcome to the portal.'),
    # Document Settings
    ('max_upload_size', 'Max Upload Size (MB)', 'document', 'number', 'Per-file size limit', '5'),
    ('allowed_file_types', 'Allowed File Types', 'document', 'text', 'Comma-separated MIME extensions', 'jpg, jpeg, png, pdf'),
    ('image_compression_quality', 'Image Compression Quality', 'document', 'number', '0-100 quality for optimized images', '80'),
    ('image_max_dimension', 'Image Max Dimension (px)', 'document', 'number', 'Longest side after resize', '2000'),
    # Security
    ('session_timeout', 'Session Timeout (minutes)', 'security', 'number', 'Idle session expiration', '30'),
    ('password_min_length', 'Min Password Length', 'security', 'number', 'Minimum password characters', '8'),
    ('login_attempts_limit', 'Login Attempts Limit', 'security', 'number', 'Lockout after failed attempts', '5'),
    ('password_policy', 'Password Policy', 'security', 'textarea', 'Password requirements description', 'Minimum 8 characters with letters and numbers.'),
    ('login_security', 'Login Security', 'security', 'textarea', 'Two-factor / captcha notes', 'Captcha required on failed attempts.'),
    # Notification Settings
    ('admin_alert_email', 'Admin Alert Email', 'notification', 'email', 'Alerts for critical events', ''),
    ('application_notifications', 'Application Notifications', 'notification', 'checkbox', 'Email applicants on status change', '1'),
]

SETTING_CATEGORIES = [
    ('application', 'Application Settings', 'description', 'Application number, autosave and submission rules.'),
    ('email', 'Email Settings', 'mail', 'SMTP configuration and outbound templates.'),
    ('document', 'Document Settings', 'folder_open', 'Upload limits, types and image compression.'),
    ('security', 'Security', 'lock', 'Session timeout and password policy.'),
    ('notification', 'Notification Settings', 'notifications', 'Admin alerts and applicant notifications.'),
]

PLATFORM_FIELDS = [
    # (key, label, field_type, help_text, choices)
    ('portal_name', 'Portal Name', 'text', 'Shown in the navbar and page titles.'),
    ('organization_name', 'Organization Name', 'text', 'Legal name of the organization.'),
    ('contact_email', 'Contact Email', 'email', 'Public support / contact email.'),
    ('contact_phone', 'Contact Phone', 'text', 'Public support phone number.'),
    ('support_email', 'Support Email', 'email', 'Helpdesk support email.'),
    ('support_phone', 'Support Phone', 'text', 'Helpdesk support phone number.'),
    ('portal_description', 'Portal Description', 'textarea', 'Short description used in SEO meta tags.'),
    ('timezone', 'Timezone', 'text', 'Default timezone, e.g. Asia/Kolkata.'),
    ('date_format', 'Date Format', 'text', 'Display format used across the portal, e.g. d M Y.'),
    ('maintenance_mode', 'Maintenance Mode', 'checkbox', 'Shows a maintenance banner across the portal.'),
    ('auto_save_enabled', 'Auto-save Drafts', 'checkbox', 'Enables draft auto-save for applicants.'),
    ('auto_save_interval', 'Auto-save Interval (seconds)', 'number', 'Draft auto-save frequency.'),
]

LOGO_ALLOWED_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.webp'}
LOGO_MAX_SIZE = 5 * 1024 * 1024
FAVICON_ALLOWED_EXTENSIONS = {'.ico', '.png'}
FAVICON_MAX_SIZE = 1 * 1024 * 1024


def _seed_setting(key, label, category, field_type, help_text, default):
    """Creates the setting row if missing (idempotent)."""
    from dashboard.models import SystemSetting
    SystemSetting.objects.get_or_create(
        key=key,
        defaults={
            'value': default,
            'category': category,
            'description': help_text,
        },
    )
    return SystemSetting.objects.get(key=key)


def _settings_context():
    """Builds the settings grouped by category with live values."""
    from dashboard.models import SystemSetting
    categories = []
    for cat_key, cat_label, icon, cat_help in SETTING_CATEGORIES:
        fields = []
        for (key, label, fcat, ftype, help_text, default) in SETTING_DEFINITIONS:
            if fcat != cat_key:
                continue
            _seed_setting(key, label, fcat, ftype, help_text, default)
            obj = SystemSetting.objects.get(key=key)
            fields.append({
                'key': key,
                'label': label,
                'field_type': ftype,
                'help_text': help_text,
                'value': obj.value,
                'description': obj.description or help_text,
            })
        categories.append({
            'key': cat_key,
            'label': cat_label,
            'icon': icon,
            'help': cat_help,
            'fields': fields,
        })
    return categories


def _system_info():
    """Read-only platform information shown on the Settings page."""
    import django as _dj
    from django.db import connection as _conn
    from dashboard.models import BackupJob, AuditLog
    db_vendor = _conn.vendor
    try:
        db_size = _conn.cursor()
        db_size.execute("SELECT page_count * page_size FROM pragma_page_count, pragma_page_size;")
        row = db_size.fetchone()
        db_bytes = row[0] if row else 0
    except Exception:
        db_bytes = 0
    media_root = os.path.join(settings.BASE_DIR, 'media')
    media_bytes = 0
    if os.path.isdir(media_root):
        for dirpath, _dirnames, filenames in os.walk(media_root):
            for fname in filenames:
                try:
                    media_bytes += os.path.getsize(os.path.join(dirpath, fname))
                except OSError:
                    pass
    last_backup = BackupJob.objects.filter(status='completed').first()
    return {
        'django_version': _dj.get_version(),
        'python_version': os.sys.version.split()[0],
        'db_engine': db_vendor,
        'db_size': db_bytes,
        'media_size': media_bytes,
        'last_backup': last_backup,
        'last_backup_at': last_backup.created_at if last_backup else None,
        'audit_log_count': AuditLog.objects.count(),
    }


@login_required
@user_passes_test(lambda u: u.is_staff)
def settings_dashboard(request):
    from dashboard.models import PlatformSetting, SystemSetting
    perm_ok = request.user.is_superuser or request.user.has_perm('dashboard.change_platform_settings')

    if request.method == 'POST':
        if not perm_ok:
            messages.error(request, 'You do not have permission to change settings.')
            return redirect('settings_dashboard')

        changed = []
        platform = PlatformSetting.get_active()

        # --- Platform identity & branding (PlatformSetting) ---
        for (key, _label, ftype, _help) in PLATFORM_FIELDS:
            raw = request.POST.get(key)
            if raw is None:
                continue
            new_value = raw if ftype == 'textarea' else raw.strip()
            if ftype == 'checkbox':
                new_value = True if raw == 'on' else False
            if ftype == 'number':
                try:
                    new_value = int(new_value)
                except (TypeError, ValueError):
                    new_value = 0
            if getattr(platform, key) != new_value:
                setattr(platform, key, new_value)
                changed.append(key)

        # Logo upload validation + save
        logo_upload = request.FILES.get('logo')
        if logo_upload:
            ext = os.path.splitext(logo_upload.name)[1].lower()
            if ext not in LOGO_ALLOWED_EXTENSIONS:
                messages.error(request, 'Logo must be a PNG, JPG, JPEG or WebP image.')
            elif logo_upload.size > LOGO_MAX_SIZE:
                messages.error(request, 'Logo must be 5 MB or smaller.')
            else:
                platform.logo = logo_upload
                changed.append('logo')

        # Favicon upload validation + save
        favicon_upload = request.FILES.get('favicon')
        if favicon_upload:
            ext = os.path.splitext(favicon_upload.name)[1].lower()
            if ext not in FAVICON_ALLOWED_EXTENSIONS:
                messages.error(request, 'Favicon must be an ICO or PNG file.')
            elif favicon_upload.size > FAVICON_MAX_SIZE:
                messages.error(request, 'Favicon must be 1 MB or smaller.')
            else:
                platform.favicon = favicon_upload
                changed.append('favicon')

        if changed:
            platform.updated_by = request.user
            platform.save()
            _write_audit_log(
                request,
                'PLATFORM_SETTINGS_CHANGED',
                f'Updated platform settings: {", ".join(changed)}.',
            )

        # --- System technical settings (SystemSetting) ---
        sys_changed = []
        for (key, _label, _fcat, _ftype, _help, _default) in SETTING_DEFINITIONS:
            raw = request.POST.get(key)
            if raw is None:
                continue
            new_value = raw if _ftype == 'textarea' else raw.strip()
            if _ftype == 'checkbox':
                new_value = '1' if raw == 'on' else ''
            obj, created = SystemSetting.objects.update_or_create(
                key=key,
                defaults={'value': new_value, 'category': _fcat, 'description': _help},
            )
            obj.updated_by = request.user
            obj.save(update_fields=['updated_by', 'updated_at', 'value'])
            sys_changed.append(key)

        if sys_changed:
            _write_audit_log(
                request,
                'SETTINGS_CHANGED',
                f'Updated settings: {", ".join(sys_changed)}.',
            )

        if changed or sys_changed:
            messages.success(request, 'Platform settings saved successfully.')
        else:
            messages.info(request, 'No settings were changed.')
        return redirect('settings_dashboard')

    platform = PlatformSetting.get_active()
    _seed_all = _settings_context()
    platform_fields = []
    for (key, label, ftype, help_text) in PLATFORM_FIELDS:
        value = getattr(platform, key, '')
        if isinstance(value, bool):
            value = '1' if value else ''
        platform_fields.append({
            'key': key,
            'label': label,
            'field_type': ftype,
            'help_text': help_text,
            'value': value,
        })
    return render(request, 'dashboard/settings.html', {
        'title': 'Platform Settings',
        'platform': platform,
        'platform_fields': platform_fields,
        'categories': _seed_all,
        'system_info': _system_info(),
        'can_change': perm_ok,
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def audit_logs(request):
    from dashboard.models import AuditLog
    from django.core.paginator import Paginator as _Paginator

    qs = AuditLog.objects.select_related('user').all()
    search = request.GET.get('q', '').strip()
    if search:
        qs = qs.filter(
            Q(action__icontains=search) |
            Q(detail__icontains=search) |
            Q(user__username__icontains=search)
        )
    paginator = _Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'dashboard/audit_logs.html', {
        'title': 'Audit Logs',
        'page_obj': page_obj,
        'q': search,
        'total': qs.count(),
    })


# ==========================================
# SYSTEM ADMINISTRATION - BACKUP & RESTORE
# ==========================================

def _human_size(num):
    """Formats a byte count into a human-readable string."""
    size = float(num or 0)
    for unit in ('B', 'KB', 'MB', 'GB'):
        if size < 1024 or unit == 'GB':
            return f"{size:.0f} {unit}" if unit == 'B' else f"{size:.1f} {unit}"
        size /= 1024
    return f"{size:.1f} GB"


def _backup_timestamp():
    return timezone.now().strftime('%Y%m%d_%H%M%S')


def _build_backup_archive(job):
    """Performs the actual backup work for a BackupJob. Returns (bytes, filename, size)."""
    import io as _io
    import zipfile
    from django.core.files.base import ContentFile

    db_path = os.path.join(settings.BASE_DIR, 'db.sqlite3')
    ts = _backup_timestamp()

    if job.backup_type == 'database':
        with open(db_path, 'rb') as fh:
            data = fh.read()
        filename = f"db_backup_{ts}.sqlite3"
        return data, filename, len(data)

    media_root = os.path.join(settings.BASE_DIR, 'media')

    def _walk_media(zf, folder=''):
        base = media_root if not folder else os.path.join(media_root, folder)
        if not os.path.isdir(base):
            return
        for dirpath, _dirnames, filenames in os.walk(base):
            for fname in filenames:
                full = os.path.join(dirpath, fname)
                # Never embed the backups folder itself.
                rel = os.path.relpath(full, media_root)
                if rel.startswith('backups') or rel.startswith('settings'):
                    continue
                try:
                    zf.write(full, arcname=os.path.join('media', rel))
                except OSError:
                    continue

    if job.backup_type == 'full':
        buffer = _io.BytesIO()
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            if os.path.exists(db_path):
                zf.write(db_path, arcname='database/db.sqlite3')
            _walk_media(zf)
        data = buffer.getvalue()
        filename = f"full_backup_{ts}.zip"
        return data, filename, len(data)

    # media-only
    buffer = _io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        _walk_media(zf)
    data = buffer.getvalue()
    filename = f"media_backup_{ts}.zip"
    return data, filename, len(data)


@login_required
@user_passes_test(lambda u: u.is_staff)
def backup_center(request):
    from dashboard.models import BackupJob
    jobs = BackupJob.objects.all()
    stats = {
        'total': jobs.count(),
        'completed': jobs.filter(status='completed').count(),
        'failed': jobs.filter(status='failed').count(),
        'total_size': sum(j.file_size for j in jobs),
    }
    return render(request, 'dashboard/backup_restore.html', {
        'title': 'Backup & Restore',
        'jobs': jobs,
        'stats': stats,
        'can_create': request.user.is_superuser or request.user.has_perm('dashboard.create_backup'),
        'can_restore': request.user.is_superuser or request.user.has_perm('dashboard.restore_backup'),
        'can_download': request.user.is_superuser or request.user.has_perm('dashboard.download_backup'),
        'human_size': _human_size,
    })


@login_required
@user_passes_test(lambda u: u.is_staff)
def backup_create(request):
    from dashboard.models import BackupJob
    from django.core.files.base import ContentFile
    from django.core.files.storage import default_storage as _store

    if not (request.user.is_superuser or request.user.has_perm('dashboard.create_backup')):
        messages.error(request, 'You do not have permission to create backups.')
        return redirect('backup_center')

    backup_type = request.POST.get('backup_type', '')
    if backup_type not in ('database', 'media', 'full'):
        messages.error(request, 'Invalid backup type selected.')
        return redirect('backup_center')

    job = BackupJob.objects.create(
        backup_type=backup_type,
        created_by=request.user,
        status='processing',
    )

    try:
        data, filename, size = _build_backup_archive(job)
        saved = _store.save(f'backups/{filename}', ContentFile(data))
        job.file.name = saved
        job.file_size = size
        job.status = 'completed'
        job.completed_at = timezone.now()
        job.save()
        _write_audit_log(request, 'BACKUP_CREATED', f'{job.backup_type_label} backup created -> {job.display_id} ({_human_size(size)}).')
        messages.success(request, f'{job.backup_type_label} backup created successfully.')
    except Exception as exc:
        job.status = 'failed'
        job.error_message = str(exc)[:2000]
        job.completed_at = timezone.now()
        job.save()
        _write_audit_log(request, 'BACKUP_FAILED', f'{job.backup_type_label} backup failed: {str(exc)[:300]}.')
        messages.error(request, f'Backup failed: {exc}')
    return redirect('backup_center')


@login_required
@user_passes_test(lambda u: u.is_staff)
def backup_download(request, job_id):
    from dashboard.models import BackupJob
    from django.http import HttpResponseForbidden
    from django.shortcuts import redirect
    if not (request.user.is_superuser or request.user.has_perm('dashboard.download_backup')):
        return HttpResponseForbidden('You do not have permission to download backups.')
    job = BackupJob.objects.filter(pk=job_id).first()
    if not job or not job.file or job.status != 'completed':
        messages.error(request, 'Backup not found or not ready for download.')
        return redirect('backup_center')
    try:
        response = FileResponse(job.file.open('rb'), content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{job.filename}"'
        _write_audit_log(request, 'BACKUP_DOWNLOADED', f'{job.backup_type_label} backup downloaded -> {job.display_id}.')
        return response
    except Exception as exc:
        messages.error(request, f'Could not open backup file: {exc}')
        return redirect('backup_center')


@login_required
@user_passes_test(lambda u: u.is_staff)
def backup_restore(request):
    """Restore flow step 1+2: select/upload backup -> show details + confirmation."""
    from dashboard.models import BackupJob
    from django.core.files.storage import default_storage as _store

    if not (request.user.is_superuser or request.user.has_perm('dashboard.restore_backup')):
        messages.error(request, 'Only authorized administrators can restore the system.')
        return redirect('backup_center')

    job = None
    preview = None

    if request.method == 'POST':
        # Option A: pick an existing backup from history
        job_id = request.POST.get('backup_id', '')
        if job_id:
            job = BackupJob.objects.filter(pk=job_id, status='completed').first()
            if not job:
                messages.error(request, 'Selected backup not found.')
                return redirect('backup_center')
            preview = {
                'display_id': job.display_id,
                'backup_type': job.backup_type_label,
                'filename': job.filename,
                'size': _human_size(job.file_size),
                'created_at': job.created_at,
                'created_by': job.created_by.username if job.created_by else 'System',
            }

        # Option B: upload a backup file
        upload = request.FILES.get('restore_file')
        if upload:
            safe_name = os.path.basename(upload.name)
            saved = _store.save(f'backups/uploads/{_backup_timestamp()}_{safe_name}', upload)
            fname_lower = safe_name.lower()
            if fname_lower.endswith('.sqlite3') or fname_lower.endswith('.db'):
                btype = 'database'
            elif 'full' in fname_lower:
                btype = 'full'
            else:
                btype = 'media'
            job = BackupJob.objects.create(
                backup_type=btype,
                created_by=request.user,
                file=saved,
                file_size=upload.size,
                status='completed',
                completed_at=timezone.now(),
            )
            preview = {
                'display_id': job.display_id,
                'backup_type': job.backup_type_label,
                'filename': job.filename,
                'size': _human_size(upload.size),
                'created_at': job.created_at,
                'created_by': request.user.username,
            }
            _write_audit_log(request, 'BACKUP_UPLOADED', f'Restore file uploaded -> {job.display_id} ({job.filename}).')

        if not preview:
            messages.error(request, 'Select a backup or upload one to continue.')
            return redirect('backup_center')

    return render(request, 'dashboard/backup_restore_confirm.html', {
        'title': 'Restore Backup',
        'preview': preview,
        'job': job,
    })


@login_required
@user_passes_test(lambda u: u.is_staff)
def backup_restore_execute(request):
    """Restore flow final step: confirmation verified -> restore -> verify -> audit log."""
    from dashboard.models import BackupJob
    from django.core.files.storage import default_storage as _store
    from django.core.files.base import ContentFile

    if not (request.user.is_superuser or request.user.has_perm('dashboard.restore_backup')):
        _write_audit_log(request, 'RESTORE_FAILED', 'Unauthorized restore attempt blocked.')
        messages.error(request, 'Only authorized administrators can restore the system.')
        return redirect('backup_center')

    job_id = request.POST.get('backup_id', '')
    confirmed = request.POST.get('confirm_restore') == 'on'
    if not confirmed:
        messages.error(request, 'You must confirm that you understand the risks before restoring.')
        return redirect('backup_restore')

    job = BackupJob.objects.filter(pk=job_id, status='completed').first()
    if not job or not job.file:
        messages.error(request, 'Backup not found.')
        return redirect('backup_center')

    _write_audit_log(request, 'RESTORE_STARTED', f'Restore started from {job.display_id} ({job.backup_type_label}).')

    try:
        from django.db import connection as _conn
        # 1) Always create a safety snapshot of the current DB first.
        db_path = os.path.join(settings.BASE_DIR, 'db.sqlite3')
        safety_ts = _backup_timestamp()
        if os.path.exists(db_path):
            with open(db_path, 'rb') as fh:
                _store.save(f'backups/safety_db_{safety_ts}.sqlite3', ContentFile(fh.read()))

        # 2) Restore per type.
        if job.backup_type in ('database', 'full'):
            _conn.close()
            with job.file.open('rb') as src:
                if job.backup_type == 'database':
                    with open(db_path, 'wb') as dst:
                        dst.write(src.read())
                else:
                    import zipfile, io as _io
                    with zipfile.ZipFile(_io.BytesIO(src.read())) as zf:
                        with zf.open('database/db.sqlite3') as srcf, open(db_path, 'wb') as dstf:
                            dstf.write(srcf.read())

        if job.backup_type in ('media', 'full'):
            import zipfile, io as _io
            media_root = os.path.join(settings.BASE_DIR, 'media')
            os.makedirs(media_root, exist_ok=True)
            with job.file.open('rb') as src:
                with zipfile.ZipFile(_io.BytesIO(src.read())) as zf:
                    for member in zf.namelist():
                        # Never restore the backups/ or settings/ folders.
                        if member.startswith('media/backups') or member.startswith('media/settings'):
                            continue
                        if member.endswith('/'):
                            continue
                        target = os.path.join(media_root, os.path.relpath(member, 'media'))
                        os.makedirs(os.path.dirname(target), exist_ok=True)
                        with zf.open(member) as srcf, open(target, 'wb') as dstf:
                            dstf.write(srcf.read())

        _write_audit_log(request, 'RESTORE_COMPLETED', f'System restored from {job.display_id} ({job.backup_type_label}).')
        messages.success(request, f'System restored from {job.display_id}. Please verify the data and re-login if needed.')
        return redirect('backup_center')
    except Exception as exc:
        _write_audit_log(request, 'RESTORE_FAILED', f'Restore from {job.display_id} failed: {str(exc)[:300]}.')
        messages.error(request, f'Restore failed: {exc}')
        return redirect('backup_center')


# ==========================================
# CUSTOM REPORTS (READ-ONLY ANALYTICS)
# ==========================================

def _report_filters(request):
    """Collect all common report filter values from the GET querystring."""
    return {
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
        'state': request.GET.get('state', ''),
        'district': request.GET.get('district', ''),
        'program': request.GET.get('program', ''),
        'qualification': request.GET.get('qualification', ''),
        'training_partner': request.GET.get('training_partner', ''),
        'application_status': request.GET.get('application_status', ''),
        'gender': request.GET.get('gender', ''),
        'category': request.GET.get('category', ''),
        'religion': request.GET.get('religion', ''),
        'occupation': request.GET.get('occupation', ''),
    }


def _apply_report_filters(qs, f):
    """Apply the common report filters (AND logic) to a queryset. Read-only."""
    if f['date_from']:
        try:
            qs = qs.filter(submission_date__date__gte=timezone.datetime.strptime(f['date_from'], '%Y-%m-%d').date())
        except ValueError:
            pass
    if f['date_to']:
        try:
            qs = qs.filter(submission_date__date__lte=timezone.datetime.strptime(f['date_to'], '%Y-%m-%d').date())
        except ValueError:
            pass
    if f['state']:
        qs = qs.filter(state_id=f['state'])
    if f['district']:
        qs = qs.filter(district_id=f['district'])
    if f['program']:
        qs = qs.filter(program_opting_id=f['program'])
    if f['qualification']:
        qs = qs.filter(qualification_id=f['qualification'])
    if f['training_partner']:
        qs = qs.filter(training_partner_id=f['training_partner'])
    if f['application_status'] == 'draft':
        qs = qs.filter(status__code='INCOMPLETE')
    elif f['application_status'] == 'submitted':
        qs = qs.filter(status__code__in=SUBMITTED_STATUS_CODES)
    elif f['application_status'] == 'pending':
        qs = qs.filter(status__code='PENDING')
    if f['gender']:
        qs = qs.filter(gender=f['gender'])
    if f['category']:
        qs = qs.filter(community_id=f['category'])
    if f['religion']:
        qs = qs.filter(religion_id=f['religion'])
    if f['occupation']:
        qs = qs.filter(occupation_id=f['occupation'])
    return qs


def _report_context_common(request):
    """Shared dropdown/master-data context used by all report pages."""
    f = _report_filters(request)
    active = {k: v for k, v in f.items() if v}
    query_string = '&'.join(f'{k}={v}' for k, v in active.items())
    return {
        'f': f,
        'query_string': query_string,
        'filters_active': bool(active),
        'states': State.objects.filter(is_active=True).order_by('name'),
        'districts': District.objects.filter(is_active=True).order_by('name'),
        'programs': Program.objects.filter(is_active=True),
        'qualifications': Qualification.objects.filter(is_active=True),
        'training_partners': TrainingPartner.objects.filter(is_active=True),
        'genders': StudentApplication.GENDER_CHOICES,
        'categories': Community.objects.filter(is_active=True),
        'religions': Religion.objects.filter(is_active=True),
        'occupations': Occupation.objects.filter(is_active=True),
        'report_statuses': [
            {'value': 'submitted', 'name': 'Submitted'},
            {'value': 'draft', 'name': 'Draft'},
            {'value': 'pending', 'name': 'Pending Review'},
        ],
    }


def _report_export_response(fmt, filename_base, title, headers, rows):
    """Generate a CSV, Excel, or PDF response from tabular report data."""
    import io as _io
    if fmt == 'excel':
        import openpyxl
        from openpyxl.styles import Font
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
        for row in rows:
            ws.append(row)
        buffer = _io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
        return response
    if fmt == 'pdf':
        from django.template.loader import render_to_string
        html = render_to_string('dashboard/report_export_pdf.html', {
            'title': title,
            'headers': headers,
            'rows': rows,
            'generated_at': timezone.now(),
        })
        buffer = _io.BytesIO()
        from xhtml2pdf import pisa
        pisa.CreatePDF(html, dest=buffer, encoding='utf-8')
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
        return response
    buffer = _io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    response = HttpResponse(buffer.getvalue().encode('utf-8'), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
    return response


@login_required
@user_passes_test(lambda u: u.is_staff)
def report_daily(request):
    """Daily Application Summary: activity grouped by submission date."""
    from django.db.models.functions import TruncDate

    f = _report_filters(request)
    qs = _apply_report_filters(StudentApplication.objects.all(), f)
    ctx = _report_context_common(request)

    total = qs.count()
    submitted = qs.filter(status__code__in=SUBMITTED_STATUS_CODES).count()
    draft = qs.filter(status__code='INCOMPLETE').count()
    pending = qs.filter(status__code='PENDING').count()
    today = timezone.localdate()
    new_today = qs.filter(submission_date__date=today).count()

    daily_rows = (
        qs.annotate(day=TruncDate('submission_date'))
        .values('day')
        .annotate(
            total=Count('id'),
            submitted=Count('id', filter=Q(status__code__in=SUBMITTED_STATUS_CODES)),
            draft=Count('id', filter=Q(status__code='INCOMPLETE')),
            pending=Count('id', filter=Q(status__code='PENDING')),
        )
        .order_by('-day')
    )
    daily_rows = list(daily_rows)

    export_rows = [
        [
            r['day'].strftime('%Y-%m-%d'),
            r['total'],
            r['submitted'],
            r['draft'],
            r['pending'],
        ]
        for r in daily_rows
    ]

    chart_days = [r['day'].strftime('%b %d, %Y') for r in reversed(daily_rows)]
    chart_totals = [r['total'] for r in reversed(daily_rows)]

    ctx.update({
        'title': 'Daily Application Summary',
        'total_applications': total,
        'submitted_count': submitted,
        'draft_count': draft,
        'pending_count': pending,
        'new_today': new_today,
        'daily_rows': daily_rows,
        'total_count': total,
        'chart_days': chart_days,
        'chart_totals': chart_totals,
        'export_url': 'report_daily_export',
        'export_title': 'Daily Application Summary',
    })
    return render(request, 'dashboard/report_daily.html', ctx)


@login_required
@user_passes_test(lambda u: u.is_staff)
def report_daily_export(request):
    from django.db.models.functions import TruncDate
    f = _report_filters(request)
    qs = _apply_report_filters(StudentApplication.objects.all(), f)
    fmt = request.GET.get('format', 'csv').lower()
    rows = (
        qs.annotate(day=TruncDate('submission_date'))
        .values('day')
        .annotate(
            total=Count('id'),
            submitted=Count('id', filter=Q(status__code__in=SUBMITTED_STATUS_CODES)),
            draft=Count('id', filter=Q(status__code='INCOMPLETE')),
            pending=Count('id', filter=Q(status__code='PENDING')),
        )
        .order_by('-day')
    )
    data = [
        [r['day'].strftime('%Y-%m-%d'), r['total'], r['submitted'], r['draft'], r['pending']]
        for r in rows
    ]
    headers = ['Date', 'Total', 'Submitted', 'Draft', 'Pending Review']
    return _report_export_response(fmt, 'daily_summary', 'Daily Application Summary', headers, data)


@login_required
@user_passes_test(lambda u: u.is_staff)
def report_demographics(request):
    """Demographic Report: distribution tables + charts over filtered applicants."""
    f = _report_filters(request)
    qs = _apply_report_filters(StudentApplication.objects.all(), f)
    ctx = _report_context_common(request)

    total = qs.count()
    male = qs.filter(gender='Male').count()
    female = qs.filter(gender='Female').count()
    other = qs.filter(gender='Other').count()
    physically_handicapped = qs.filter(physically_handicapped='Yes').count()
    ex_serviceman = qs.filter(ex_serviceman__name='Yes').count()

    def distribution(qs, field):
        """Return [(label, count, pct)] for a values/annotate field (read-only)."""
        rows = list(qs.values(field).annotate(count=Count('id')).order_by('-count'))
        out = []
        for r in rows:
            label = r[field]
            if not label:
                label = 'Not Specified'
            pct = round((r['count'] / total * 100), 1) if total else 0
            out.append({'label': label, 'count': r['count'], 'percentage': pct})
        return out

    gender_dist = distribution(qs, 'gender')
    category_dist = distribution(qs, 'community__name')
    religion_dist = distribution(qs, 'religion__name')
    marital_dist = distribution(qs, 'marital_status__name')
    nationality_dist = distribution(qs, 'nationality')
    occupation_dist = distribution(qs, 'occupation__name')

    def export_rows(dist, key):
        return [[r['label'], r['count'], f"{r['percentage']}%"] for r in dist]

    export_rows_data = []
    for section, dist in [
        ('Gender', gender_dist),
        ('Category', category_dist),
        ('Religion', religion_dist),
        ('Marital Status', marital_dist),
        ('Nationality', nationality_dist),
        ('Occupation', occupation_dist),
    ]:
        for r in dist:
            export_rows_data.append([section, r['label'], r['count'], f"{r['percentage']}%"])

    ctx.update({
        'title': 'Demographic Report',
        'total_students': total,
        'male': male,
        'female': female,
        'other': other,
        'physically_handicapped': physically_handicapped,
        'ex_serviceman': ex_serviceman,
        'total_count': total,
        'gender_dist': gender_dist,
        'category_dist': category_dist,
        'religion_dist': religion_dist,
        'marital_dist': marital_dist,
        'nationality_dist': nationality_dist,
        'occupation_dist': occupation_dist,
        'gender_labels': [r['label'] for r in gender_dist],
        'gender_data': [r['count'] for r in gender_dist],
        'category_labels': [r['label'] for r in category_dist],
        'category_data': [r['count'] for r in category_dist],
        'religion_labels': [r['label'] for r in religion_dist],
        'religion_data': [r['count'] for r in religion_dist],
        'export_url': 'report_demographics_export',
        'export_title': 'Demographic Report',
        'export_headers': ['Section', 'Value', 'Count', 'Percentage'],
        'export_rows': export_rows_data,
    })
    return render(request, 'dashboard/report_demographics.html', ctx)


@login_required
@user_passes_test(lambda u: u.is_staff)
def report_demographics_export(request):
    f = _report_filters(request)
    qs = _apply_report_filters(StudentApplication.objects.all(), f)
    fmt = request.GET.get('format', 'csv').lower()

    total = qs.count()

    def rows_for(field):
        out = []
        for r in qs.values(field).annotate(count=Count('id')).order_by('-count'):
            label = r[field] or 'Not Specified'
            pct = round((r['count'] / total * 100), 1) if total else 0
            out.append((label, r['count'], f"{pct}%"))
        return out

    data = []
    for section, field in [
        ('Gender', 'gender'),
        ('Category', 'community__name'),
        ('Religion', 'religion__name'),
        ('Marital Status', 'marital_status__name'),
        ('Nationality', 'nationality'),
        ('Occupation', 'occupation__name'),
    ]:
        for label, count, pct in rows_for(field):
            data.append([section, label, count, pct])

    headers = ['Section', 'Value', 'Count', 'Percentage']
    return _report_export_response(fmt, 'demographics', 'Demographic Report', headers, data)


@login_required
@user_passes_test(lambda u: u.is_staff)
def report_geographical(request):
    """Geographical Report: state & district-wise application analysis."""
    f = _report_filters(request)
    qs = _apply_report_filters(StudentApplication.objects.all(), f)
    ctx = _report_context_common(request)

    total = qs.count()
    total_states = qs.exclude(state__isnull=True).values('state').distinct().count()
    total_districts = qs.exclude(district__isnull=True).values('district').distinct().count()

    geo_rows = (
        qs.exclude(state__isnull=True, district__isnull=True)
        .values('state__name', 'district__name')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    geo_rows = list(geo_rows)

    state_totals = (
        qs.exclude(state__isnull=True)
        .values('state__name')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    state_totals = list(state_totals)
    district_totals = (
        qs.exclude(district__isnull=True)
        .values('district__name')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    district_totals = list(district_totals)

    top_state = state_totals[0]['state__name'] if state_totals else '—'
    top_district = district_totals[0]['district__name'] if district_totals else '—'

    for r in geo_rows:
        r['percentage'] = round((r['count'] / total * 100), 1) if total else 0

    export_rows = [
        [r['state__name'] or '—', r['district__name'] or '—', r['count'], f"{r['percentage']}%"]
        for r in geo_rows
    ]

    ctx.update({
        'title': 'Geographical Report',
        'subtitle': 'State & District-wise Application Analysis',
        'total_states': total_states,
        'total_districts': total_districts,
        'total_applications': total,
        'top_state': top_state,
        'top_district': top_district,
        'geo_rows': geo_rows,
        'total_count': total,
        'chart_states': [r['state__name'] for r in state_totals],
        'chart_state_counts': [r['count'] for r in state_totals],
        'chart_districts': [r['district__name'] for r in district_totals[:15]],
        'chart_district_counts': [r['count'] for r in district_totals[:15]],
        'export_url': 'report_geographical_export',
        'export_title': 'Geographical Report',
        'export_headers': ['State', 'District', 'Applications', 'Percentage'],
        'export_rows': export_rows,
    })
    return render(request, 'dashboard/report_geographical.html', ctx)


@login_required
@user_passes_test(lambda u: u.is_staff)
def report_geographical_export(request):
    f = _report_filters(request)
    qs = _apply_report_filters(StudentApplication.objects.all(), f)
    fmt = request.GET.get('format', 'csv').lower()

    total = qs.count()
    geo_rows = (
        qs.exclude(state__isnull=True, district__isnull=True)
        .values('state__name', 'district__name')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    data = []
    for r in geo_rows:
        pct = round((r['count'] / total * 100), 1) if total else 0
        data.append([r['state__name'] or '—', r['district__name'] or '—', r['count'], f"{pct}%"])

    headers = ['State', 'District', 'Applications', 'Percentage']
    return _report_export_response(fmt, 'geographical', 'Geographical Report', headers, data)


import csv
import os
from django.http import HttpResponse, FileResponse
from django.conf import settings

@login_required
@user_passes_test(lambda u: u.is_staff)
def export_csv(request):
    """
    Generates a live CSV export of the currently-filtered student applications.
    Respects the same filters as the Applications page (search, state, district,
    program, qualification, partner, status, gender, category, occupation, year,
    submission date). Uses the full comprehensive export logic.
    """
    from utilities.export_generator import generate_full_csv_export
    from django.http import HttpResponse
    
    applications = _filtered_applications(request)
    csv_data = generate_full_csv_export(applications, request)
    
    response = HttpResponse(csv_data, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="applications_export.csv"'
    return response

@login_required
@user_passes_test(lambda u: u.is_staff)
def export_db(request):
    """
    Securely downloads the live db.sqlite3 database backup.
    """
    db_path = os.path.join(settings.BASE_DIR, 'db.sqlite3')
    if os.path.exists(db_path):
        response = FileResponse(open(db_path, 'rb'), content_type='application/x-sqlite3')
        response['Content-Disposition'] = 'attachment; filename="backup_db.sqlite3"'
        return response
    return HttpResponse("Database file not found.", status=404)


def _collect_selected_ids(request):
    """
    Reads and validates selected application IDs from a POST body.
    Returns a list of valid positive integer PKs only. Invalid/non-numeric
    values are silently ignored (never trusted from the browser).
    """
    ids = request.POST.getlist('selected_ids[]')
    valid = []
    for raw in ids:
        try:
            pk = int(raw)
        except (TypeError, ValueError):
            continue
        if pk > 0:
            valid.append(pk)
    return valid


def _selected_applications_queryset(request):
    """
    Resolves validated selected IDs into a real queryset.
    Only existing records are kept; non-existent IDs are ignored.
    """
    from registrations.models import StudentApplication
    ids = _collect_selected_ids(request)
    if not ids:
        return StudentApplication.objects.none()
    return StudentApplication.objects.filter(id__in=ids)


def _bulk_timestamp():
    """YYYYMMDD_HHMMSS timestamp used in CSC bulk download filenames."""
    return timezone.now().strftime('%Y%m%d_%H%M%S')


def _write_audit_log(request, action, detail='', app_ids=None):
    """Writes an AuditLog entry for a bulk operation (best-effort, never raises)."""
    from dashboard.models import AuditLog
    try:
        AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action=action,
            detail=detail[:2000],
            application_ids=','.join(str(i) for i in (app_ids or [])),
        )
    except Exception:
        pass


def _bulk_export_data(queryset, fmt, request):
    """Shared CSV/Excel generation used by dedicated export endpoints."""
    from utilities.export_generator import generate_full_csv_export, generate_full_excel_export
    if fmt == 'excel':
        data = generate_full_excel_export(queryset, request)
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = f"CSC_Selected_Students_{_bulk_timestamp()}.xlsx"
    else:
        data = generate_full_csv_export(queryset, request)
        content_type = 'text/csv; charset=utf-8'
        filename = f"CSC_Selected_Students_{_bulk_timestamp()}.csv"
    return data, content_type, filename


@login_required
@user_passes_test(lambda u: u.is_staff)
def bulk_export(request):
    """
    Legacy combined endpoint: exports ONLY the selected applications to CSV or Excel.
    Format is chosen via POST field 'format' ('csv' default, 'excel' supported).
    Kept for backward compatibility; dedicated endpoints are bulk_export_csv/excel.
    """
    from django.http import HttpResponse, JsonResponse

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed.'}, status=405)

    ids = _collect_selected_ids(request)
    if not ids:
        return JsonResponse({'error': 'Please select at least one student.'}, status=400)

    queryset = _selected_applications_queryset(request)
    fmt = request.POST.get('format', 'csv').lower()
    try:
        data, content_type, filename = _bulk_export_data(queryset, fmt, request)
    except Exception as exc:
        return JsonResponse({'error': f'Export failed: {exc}'}, status=500)

    response = HttpResponse(data, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@user_passes_test(lambda u: u.is_staff)
def bulk_export_csv(request):
    """Exports ONLY the selected applications to a CSC_Selected_Students_*.csv file."""
    from django.http import HttpResponse, JsonResponse

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed.'}, status=405)

    ids = _collect_selected_ids(request)
    if not ids:
        return JsonResponse({'error': 'Please select at least one student.'}, status=400)

    queryset = _selected_applications_queryset(request)
    try:
        data, content_type, filename = _bulk_export_data(queryset, 'csv', request)
    except Exception as exc:
        return JsonResponse({'error': f'Export failed: {exc}'}, status=500)

    response = HttpResponse(data, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@user_passes_test(lambda u: u.is_staff)
def bulk_export_excel(request):
    """Exports ONLY the selected applications to a CSC_Selected_Students_*.xlsx file."""
    from django.http import HttpResponse, JsonResponse

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed.'}, status=405)

    ids = _collect_selected_ids(request)
    if not ids:
        return JsonResponse({'error': 'Please select at least one student.'}, status=400)

    queryset = _selected_applications_queryset(request)
    try:
        data, content_type, filename = _bulk_export_data(queryset, 'excel', request)
    except Exception as exc:
        return JsonResponse({'error': f'Export failed: {exc}'}, status=500)

    response = HttpResponse(data, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@user_passes_test(lambda u: u.is_staff)
def bulk_download_documents(request):
    """
    Creates a CSC_Selected_Documents_*.zip containing documents belonging ONLY
    to the selected applications, structured as CSC_Selected_Documents/APPLICATION_NO/<type>/.
    Missing documents are skipped; safe sanitized archive names are used.
    """
    import zipfile
    import io
    import os
    import re
    from django.http import HttpResponse, JsonResponse

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed.'}, status=405)

    def safe_name(value):
        name = re.sub(r'[^\w\- ]', '_', str(value or ''))
        name = re.sub(r'[\s]+', '_', name)
        name = name.strip('._')
        return name or 'Application'

    doc_folders = {
        'passport_photo': 'Passport_Photo',
        'signature': 'Signature',
        'thumb_impression': 'Left_Thumb',
        'aadhaar_pdf': 'Aadhaar',
        'aadhaar': 'Aadhaar',
        'community_certificate': 'Community_Certificate',
        'community_certificate_doc': 'Community_Certificate',
        'registration_screenshot': 'Registration_Screenshot',
        'abc_screenshot': 'ABC_ID',
        'supporting_documents': 'Supporting_Documents',
        'additional_documents': 'Additional_Documents',
    }

    ids = _collect_selected_ids(request)
    if not ids:
        return JsonResponse({'error': 'Please select at least one student.'}, status=400)

    queryset = _selected_applications_queryset(request).prefetch_related('documents')

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for app in queryset:
            app_no = safe_name(app.application_number) if app.application_number else f"App_{app.id}"
            app_folder = f"CSC_Selected_Documents/{app_no}/"

            if app.acknowledgement_pdf and app.acknowledgement_pdf.name:
                try:
                    if os.path.exists(app.acknowledgement_pdf.path):
                        ext = os.path.splitext(app.acknowledgement_pdf.name)[1] or '.pdf'
                        zip_file.write(app.acknowledgement_pdf.path, arcname=f"{app_folder}Acknowledgement{ext}")
                except Exception:
                    pass

            for doc in app.documents.all():
                folder_name = doc_folders.get(doc.doc_type, safe_name(doc.doc_type))
                for field in ('original_file', 'compressed_file'):
                    file_field = getattr(doc, field, None)
                    if file_field and file_field.name:
                        try:
                            if os.path.exists(file_field.path):
                                ext = os.path.splitext(file_field.name)[1] or ''
                                suffix = 'optimized' if field == 'compressed_file' else 'original'
                                zip_file.write(file_field.path, arcname=f"{app_folder}{folder_name}/{suffix}{ext}")
                        except Exception:
                            continue

    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="CSC_Selected_Documents_{_bulk_timestamp()}.zip"'
    return response


@login_required
@user_passes_test(lambda u: u.is_staff)
def bulk_correction(request):
    """
    Marks the selected SUBMITTED applications for correction (verification_status=CORRECTION).
    Draft/incomplete records are skipped because the correction workflow only applies
    to officially submitted applications. No new status system is created.
    """
    from django.db import transaction
    from django.http import JsonResponse
    from masterdata.models import VerificationStatus

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed.'}, status=405)

    ids = _collect_selected_ids(request)
    if not ids:
        return JsonResponse({'error': 'Please select at least one student.'}, status=400)

    correction_status = VerificationStatus.objects.filter(code='CORRECTION').first()
    if not correction_status:
        return JsonResponse({'error': 'CORRECTION verification status not found.'}, status=400)

    queryset = _selected_applications_queryset(request)
    total = queryset.count()
    with transaction.atomic():
        processed = queryset.exclude(status__code='INCOMPLETE').update(verification_status=correction_status)
    skipped = total - processed
    if processed:
        _write_audit_log(request, 'BULK_CORRECTION', f'Marked {processed} application(s) for correction.', ids)

    return JsonResponse({'processed': processed, 'skipped': skipped})


@login_required
@user_passes_test(lambda u: u.is_staff)
def bulk_request_correction(request):
    """
    Dedicated endpoint: requests correction for ONLY the selected applications.
    Same workflow as bulk_correction but exposed under the bulk-operations URL space.
    """
    return bulk_correction(request)


@login_required
@user_passes_test(lambda u: u.is_staff)
def bulk_verify(request):
    """
    Marks the selected SUBMITTED applications as Verified (verification_status=VERIFIED).
    Draft/incomplete records are skipped; the application status is never changed
    (verification does NOT auto-approve or reject anything).
    """
    from django.db import transaction
    from django.http import JsonResponse
    from masterdata.models import VerificationStatus

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed.'}, status=405)

    ids = _collect_selected_ids(request)
    if not ids:
        return JsonResponse({'error': 'Please select at least one student.'}, status=400)

    verified_status = VerificationStatus.objects.filter(code='VERIFIED').first()
    if not verified_status:
        return JsonResponse({'error': 'VERIFIED verification status not found.'}, status=400)

    queryset = _selected_applications_queryset(request)
    total = queryset.count()
    with transaction.atomic():
        processed = queryset.exclude(status__code='INCOMPLETE').update(verification_status=verified_status)
    skipped = total - processed
    if processed:
        _write_audit_log(request, 'BULK_VERIFY', f'Verified {processed} application(s).', ids)

    return JsonResponse({'processed': processed, 'skipped': skipped})

