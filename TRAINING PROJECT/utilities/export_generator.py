import csv
import io
import openpyxl
from openpyxl.utils import get_column_letter
from django.conf import settings

# Comprehensive list of columns expected by the user prompt
EXPORT_COLUMNS = [
    # 1. Personal Details
    "Name (as per Aadhaar)", "Father's Name", "Mother's Name", "Date of Birth",
    "Gender", "Religion", "Community", "Marital Status", "Ex-Service Man",
    "Category", "Nationality", "Physically Handicapped", "Annual Income",
    
    # 2. Contact Details
    "Mobile Number", "Aadhaar Number", "Alternative Mobile", "Email Address",
    "Full Address", "Country", "State", "District", "Pincode", "Occupation",
    
    # 3. Educational Details
    "Applying Qualification", "Name of the program opting", "Mode of Qualification",
    "Stream of Qualification", "Completion Status", "Year of Passing", "Percentage (%)",
    "Name of Educational Institution", "Training Partner/ATC", "Batch Code",
    "Hall Ticket Number", "Registration Number", "ABC ID (Apaar ID)",
    
    # 4. Identification
    "Distinguishing Mark",
    
    # 5. Document Uploads (Dynamically added below)
    # We will generate 4 columns for each document: Name, URL, Type, Size
]

DOC_TYPES = {
    'Passport_Photo': 'Upload Photo',
    'Signature': 'Upload Signature',
    'Left_Thumb': 'Upload Left Hand Thumb Impression',
    'Father_Signature': "Upload Father's/Guardian's Signature",
    'Aadhaar': 'Aadhaar Document',
    'Community_certificate': 'Community Certificate',
    'Additional_Documents': 'Additional Documents (Single PDF)',
    # Educational certificates are combined in Additional Documents based on user instructions
}

for doc_key, doc_label in DOC_TYPES.items():
    EXPORT_COLUMNS.extend([
        f"{doc_label}", f"{doc_label} File Type", f"{doc_label} File Size"
    ])

# 6. Declaration & 7. Application Information
EXPORT_COLUMNS.extend([
    "Declaration / Consent", "Declaration Accepted Date",
    "Application Number", "Application Status", "Verification Status", 
    "Registration Date", "Last Updated Date", "IP Address", "Browser Info", "Remarks"
])


def extract_application_row(app, request=None):
    """
    Extracts all fields from a StudentApplication instance matching EXPORT_COLUMNS exactly.
    """
    row = []

    # Safe getattr for foreign keys
    def f_name(fk):
        return fk.name if fk else ""

    # 1. Personal Details
    row.extend([
        app.full_name or "",
        app.father_name or "",
        app.mother_name or "",
        app.dob.strftime('%Y-%m-%d') if app.dob else "",
        app.gender or "",
        f_name(app.religion),
        f_name(app.community),
        f_name(app.marital_status),
        f_name(app.ex_serviceman),
        app.category or "",
        app.nationality or "",
        app.physically_handicapped or "",
        str(app.annual_income) if app.annual_income is not None else "",
    ])
    
    # 2. Contact Details
    # Prefix text IDs with single quote to prevent Excel dropping leading zeros or scientific notation
    row.extend([
        f"'{app.mobile_number}" if app.mobile_number else "",
        f"'{app.aadhaar_number}" if app.aadhaar_number else "",
        f"'{app.alternative_mobile}" if app.alternative_mobile else "",
        app.email or "",
        app.communication_address or "",
        f_name(app.country),
        f_name(app.state),
        f_name(app.district),
        f"'{app.pincode}" if app.pincode else "",
        f_name(app.occupation),
    ])
    
    # 3. Educational Details
    stream = app.custom_stream if app.custom_stream else f_name(app.qualification)
    row.extend([
        app.applying_qualification or "",
        f_name(app.program_opting),
        app.mode_of_qualification or "",
        stream,
        app.completion_status or "",
        str(app.year_of_passing) if app.year_of_passing else "",
        str(app.percentage) if app.percentage else "",
        app.institution or "",
        app.training_partner or "",
        app.batch_code or "",
        f"'{app.hall_ticket_number}" if app.hall_ticket_number else "",
        f"'{app.registration_number}" if app.registration_number else "",
        f"'{app.abc_id}" if app.abc_id else "",
    ])
    
    # 4. Identification
    row.extend([
        app.distinguishing_mark or "",
    ])
    
    # 5. Documents
    docs = {doc.doc_type: doc for doc in app.documents.all()}
    base_url = request.build_absolute_uri('/')[:-1] if request else settings.SITE_URL if hasattr(settings, 'SITE_URL') else ""
    
    for doc_key, doc_label in DOC_TYPES.items():
        doc = docs.get(doc_key)
        if doc and (doc.compressed_file or doc.original_file):
            file_obj = doc.compressed_file if doc.compressed_file else doc.original_file
            
            # File Name & URL
            file_url = base_url + file_obj.url if file_obj else ""
            
            # Type
            file_type = "PDF" if doc.mime_type == 'application/pdf' else doc.mime_type
            
            # Size
            size = f"{doc.compressed_size_kb} KB" if doc.compressed_file else (f"{doc.original_size_kb} KB" if doc.original_size_kb else "")
            
            row.extend([file_url, file_type or "", size])
        else:
            row.extend(["", "", ""])
            
    # 6. Declaration
    row.extend([
        "Accepted" if getattr(app, 'declaration_accepted', False) else "Not Accepted",
        app.submission_date.strftime('%Y-%m-%d %H:%M:%S') if app.submission_date else "",
    ])
    
    # 7. Application Information
    row.extend([
        app.application_number or "",
        f_name(app.status),
        f_name(app.verification_status),
        app.submission_date.strftime('%Y-%m-%d %H:%M:%S') if app.submission_date else "",
        app.last_updated.strftime('%Y-%m-%d %H:%M:%S') if app.last_updated else "",
        app.ip_address or "",
        app.browser_info or "",
        app.remarks or "",
    ])
    
    return row

def generate_full_excel_export(queryset, request=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"
    
    # Header
    ws.append(EXPORT_COLUMNS)
    
    # Styling Header
    from openpyxl.styles import Font
    for col in range(1, len(EXPORT_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        # Set column width
        ws.column_dimensions[get_column_letter(col)].width = 25
        
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    
    # Data
    # Pre-fetch for optimization
    queryset = queryset.select_related(
        'religion', 'community', 'marital_status', 'ex_serviceman', 'occupation',
        'country', 'state', 'district', 'program_opting', 'qualification',
        'status', 'verification_status'
    ).prefetch_related('documents')
    
    for app in queryset:
        ws.append(extract_application_row(app, request))
        # Make URLs clickable hyperlinks
        current_row = ws.max_row
        for col in range(1, len(EXPORT_COLUMNS) + 1):
            cell = ws.cell(row=current_row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith('http'):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

def generate_full_csv_export(queryset, request=None):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    
    writer.writerow(EXPORT_COLUMNS)
    
    queryset = queryset.select_related(
        'religion', 'community', 'marital_status', 'ex_serviceman', 'occupation',
        'country', 'state', 'district', 'program_opting', 'qualification',
        'status', 'verification_status'
    ).prefetch_related('documents')
    
    for app in queryset:
        writer.writerow(extract_application_row(app, request))
        
    return buffer.getvalue().encode('utf-8')
